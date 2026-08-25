import ast
import importlib
import inspect
from concurrent.futures import ThreadPoolExecutor
from dataclasses import FrozenInstanceError
from datetime import date, datetime, timedelta
from decimal import Decimal
from threading import Barrier, RLock
from types import MappingProxyType
from typing import get_type_hints
from uuid import UUID

import pytest
from django.core.exceptions import ValidationError
from django.utils import timezone


@pytest.fixture
def isolated_report_registry(monkeypatch):
    import reports.registry as registry

    executors = {}
    monkeypatch.setattr(registry, '_EXECUTORS', executors)
    monkeypatch.setattr(registry, '_EXECUTOR_LOCK', RLock(), raising=False)
    return registry, executors


def test_report_contracts_are_immutable_and_dataset_keeps_rows_lazy():
    from reports.contracts import ReportColumn, ReportDataset

    consumed = False

    def rows():
        nonlocal consumed
        consumed = True
        yield {'amount': Decimal('10.0000')}

    column = ReportColumn('amount', 'Valor', 'decimal')
    dataset = ReportDataset(
        title='Exemplo',
        columns=(column,),
        rows=rows(),
    )

    assert consumed is False
    assert list(dataset.rows) == [{'amount': Decimal('10.0000')}]
    assert consumed is True
    with pytest.raises(FrozenInstanceError):
        column.label = 'Total'


def test_report_registry_resolves_only_registered_normalized_keys(isolated_report_registry):
    from reports.contracts import ReportColumn, ReportDataset

    registry, _executors = isolated_report_registry

    @registry.register_executor('tests.task1.example')
    def example(context):
        return ReportDataset(
            title='Exemplo',
            columns=(ReportColumn('amount', 'Valor', 'decimal'),),
            rows=({'amount': Decimal('10.0000')},),
        )

    assert registry.get_executor('tests.task1.example') is example

    with pytest.raises(ValidationError) as unknown:
        registry.get_executor('tests.task1.unknown')

    assert unknown.value.message_dict == {
        'executor_key': ['Executor de relatório não registrado.'],
    }


@pytest.mark.parametrize(
    'key',
    [
        '',
        'single',
        'finance..cash_flow',
        '.finance.cash_flow',
        'finance.cash_flow.',
        'Finance.cash_flow',
        'finance.CashFlow',
        'finance/cash_flow',
        '../finance.cash_flow',
        'finance.cash flow',
        'finance.cash_flow;drop',
        'finance.cash_flow\nother',
        123,
    ],
)
def test_report_registry_rejects_noncanonical_keys_for_register_and_get(
    isolated_report_registry,
    key,
):
    registry, _executors = isolated_report_registry

    with pytest.raises(RuntimeError, match='Chave de executor de relatório inválida'):
        registry.register_executor(key)

    with pytest.raises(ValidationError) as invalid:
        registry.get_executor(key)

    assert invalid.value.message_dict == {
        'executor_key': ['Executor de relatório não registrado.'],
    }


def test_report_registry_accepts_canonical_dotted_key(isolated_report_registry):
    registry, _executors = isolated_report_registry

    @registry.register_executor('finance.cash_flow')
    def cash_flow(context):
        return context

    assert registry.get_executor('finance.cash_flow') is cash_flow


def test_report_registry_rejects_duplicate_keys_deterministically(isolated_report_registry):
    registry, _executors = isolated_report_registry

    key = 'tests.task1.duplicate'

    @registry.register_executor(key)
    def first(context):
        return context

    with pytest.raises(
        RuntimeError,
        match=r'^Executor de relatório duplicado: tests\.task1\.duplicate$',
    ):

        @registry.register_executor(key)
        def second(context):
            return context


def test_report_registry_registers_exactly_one_executor_during_thread_race(
    isolated_report_registry,
):
    registry, _executors = isolated_report_registry
    worker_count = 8
    barrier = Barrier(worker_count)

    def attempt(index):
        def executor(context):
            return context

        barrier.wait()
        try:
            registry.register_executor('tests.task1.concurrent')(executor)
        except RuntimeError as exc:
            return 'duplicate', str(exc), index
        return 'registered', executor, index

    with ThreadPoolExecutor(max_workers=worker_count) as pool:
        results = list(pool.map(attempt, range(worker_count)))

    registered = [result for result in results if result[0] == 'registered']
    duplicates = [result for result in results if result[0] == 'duplicate']
    assert len(registered) == 1
    assert len(duplicates) == worker_count - 1
    assert {result[1] for result in duplicates} == {
        'Executor de relatório duplicado: tests.task1.concurrent'
    }
    assert registry.get_executor('tests.task1.concurrent') is registered[0][1]


def test_report_registry_uses_a_dedicated_lock():
    import reports.registry as registry

    assert registry._EXECUTOR_LOCK is not None


def test_report_registry_rejects_string_subclass_without_invoking_dunders(
    isolated_report_registry,
):
    registry, _executors = isolated_report_registry

    class ExecutorKey(str):
        pass

    key = ExecutorKey('finance.cash_flow')

    def explode(*args):
        raise AssertionError('dunder da chave não pode ser chamado')

    ExecutorKey.__bool__ = explode
    ExecutorKey.__eq__ = explode
    ExecutorKey.__hash__ = explode
    ExecutorKey.__str__ = explode

    with pytest.raises(RuntimeError, match='Chave de executor de relatório inválida'):
        registry.register_executor(key)

    with pytest.raises(ValidationError):
        registry.get_executor(key)


def test_reports_ready_and_executor_import_are_idempotent(isolated_report_registry):
    from reports.apps import ReportsConfig

    registry, executors = isolated_report_registry

    @registry.register_executor('tests.task1.sentinel')
    def sentinel(context):
        return context

    config = ReportsConfig('reports', importlib.import_module('reports'))
    first_package = importlib.import_module('reports.executors')
    config.ready()
    config.ready()
    second_package = importlib.import_module('reports.executors')

    assert first_package is second_package
    assert executors == {'tests.task1.sentinel': sentinel}


def test_registry_fixture_does_not_leak_registered_executors(isolated_report_registry):
    registry, executors = isolated_report_registry

    assert executors == {}
    with pytest.raises(ValidationError):
        registry.get_executor('tests.task1.sentinel')


def test_report_executors_package_can_be_imported_before_curated_modules_exist():
    from reports import executors

    assert executors is not None


def test_filter_normalization_rejects_unknown_and_inverted_period():
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as unknown:
        normalize_report_filters({'sql': 'select 1'}, allowed=('period_start',))
    assert unknown.value.message_dict == {
        'filters': ['Filtros não suportados: sql.'],
    }

    with pytest.raises(ValidationError) as inverted:
        normalize_report_filters(
            {'period_start': '2026-07-31', 'period_end': '2026-07-01'},
            allowed=('period_start', 'period_end'),
        )
    assert inverted.value.message_dict == {
        'period_end': ['Data final não pode ser anterior à inicial.'],
    }

    result = normalize_report_filters(
        {'period_start': '2026-07-01'},
        allowed=('period_start',),
    )
    assert result['period_start'] == date(2026, 7, 1)


def test_filter_normalization_rejects_inverted_due_period():
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as inverted:
        normalize_report_filters(
            {'due_start': '2026-08-01', 'due_end': '2026-07-31'},
            allowed=('due_start', 'due_end'),
        )

    assert inverted.value.message_dict == {
        'due_end': ['Data final não pode ser anterior à inicial.'],
    }


def test_filter_normalization_treats_none_as_empty_mapping():
    from reports.filters import normalize_report_filters

    assert normalize_report_filters(None, allowed=('status',)) == {}


def test_normalized_filters_compose_directly_with_report_context():
    from django.contrib.auth import get_user_model

    from reports.contracts import ReportContext, ReportFilters
    from reports.filters import normalize_report_filters

    normalized = normalize_report_filters(
        {'status': 'active'},
        allowed=('status',),
    )
    context = ReportContext(filters=normalized, user=get_user_model()())

    assert get_type_hints(ReportContext)['filters'] == ReportFilters
    assert normalize_report_filters(
        context.filters,
        allowed=('status',),
    ) == {'status': 'active'}


def test_filter_normalization_rejects_read_only_mapping():
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(
            MappingProxyType({'status': 'active'}),
            allowed=('status',),
        )

    assert invalid.value.message_dict == {
        'filters': ['Filtros devem ser um objeto.'],
    }


@pytest.mark.parametrize('filters', [[], 'status=active', object()])
def test_filter_normalization_accepts_only_dict_objects(filters):
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(filters, allowed=('status',))

    assert invalid.value.message_dict == {
        'filters': ['Filtros devem ser um objeto.'],
    }


def test_filter_normalization_rejects_dict_subclass_without_iterating_or_testing_truth():
    from reports.filters import normalize_report_filters

    class ExplosiveDict(dict):
        def __bool__(self):
            raise AssertionError('__bool__ não pode ser chamado')

        def __iter__(self):
            raise AssertionError('__iter__ não pode ser chamado')

        def items(self):
            raise AssertionError('items não pode ser chamado')

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(ExplosiveDict(), allowed=('status',))

    assert invalid.value.message_dict == {
        'filters': ['Filtros devem ser um objeto.'],
    }


def test_filter_normalization_rejects_non_exact_string_keys_before_membership_operations():
    from reports.filters import normalize_report_filters

    class StringKey(str):
        pass

    key = StringKey('status')
    filters = {key: 'active'}

    def explode(*args):
        raise AssertionError('dunder da chave não pode ser chamado')

    StringKey.__eq__ = explode
    StringKey.__hash__ = explode
    StringKey.__str__ = explode

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(filters, allowed=('status',), required=('status',))

    assert invalid.value.message_dict == {
        'filters': ['As chaves dos filtros devem ser textos.'],
    }


@pytest.mark.parametrize('argument_name', ['allowed', 'required'])
def test_filter_normalization_rejects_non_exact_internal_filter_keys(argument_name):
    from reports.filters import normalize_report_filters

    class StringKey(str):
        pass

    key = StringKey('status')

    def explode(*args):
        raise AssertionError('dunder da configuração não pode ser chamado')

    StringKey.__bool__ = explode
    StringKey.__eq__ = explode
    StringKey.__hash__ = explode
    StringKey.__str__ = explode

    kwargs = {'allowed': ('status',), 'required': ()}
    kwargs[argument_name] = (key,)

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters({'status': 'active'}, **kwargs)

    assert invalid.value.message_dict == {
        argument_name: ['As chaves de filtros configuradas devem ser textos.'],
    }


@pytest.mark.parametrize(
    'value',
    [
        {'nested': 'value'},
        ['active'],
        ('active',),
        {'active'},
        Decimal('1.00'),
        object(),
    ],
)
def test_filter_normalization_rejects_non_primitive_values(value):
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters({'status': value}, allowed=('status',))

    assert invalid.value.message_dict == {
        'status': ['Informe um valor de filtro primitivo.'],
    }


def test_filter_normalization_does_not_evaluate_complex_filter_objects():
    from reports.filters import normalize_report_filters

    class QueryFragment:
        def __eq__(self, other):
            raise AssertionError('objetos complexos não podem ser avaliados')

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(
            {'status': QueryFragment()},
            allowed=('status',),
            required=('status',),
        )

    assert invalid.value.message_dict == {
        'status': ['Informe um valor de filtro primitivo.'],
    }


def test_filter_normalization_never_invokes_untrusted_value_dunders():
    from reports.filters import normalize_report_filters

    class ExplosiveValue:
        def __bool__(self):
            raise AssertionError('__bool__ não pode ser chamado')

        def __eq__(self, other):
            raise AssertionError('__eq__ não pode ser chamado')

        def __iter__(self):
            raise AssertionError('__iter__ não pode ser chamado')

        def __hash__(self):
            raise AssertionError('__hash__ não pode ser chamado')

        def __str__(self):
            raise AssertionError('__str__ não pode ser chamado')

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(
            {'status': ExplosiveValue()},
            allowed=('status',),
            required=('status',),
        )

    assert invalid.value.message_dict == {
        'status': ['Informe um valor de filtro primitivo.'],
    }


@pytest.mark.parametrize(
    'value',
    [
        type('StringValue', (str,), {})('active'),
        type('IntegerValue', (int,), {})(1),
        type('FloatValue', (float,), {})(1.5),
        type('DateValue', (date,), {})(2026, 7, 31),
    ],
)
def test_filter_normalization_rejects_scalar_subclasses(value):
    from reports.filters import normalize_report_filters

    key = 'period_start' if isinstance(value, date) else 'status'
    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters({key: value}, allowed=(key,))

    assert invalid.value.message_dict == {
        key: ['Informe um valor de filtro primitivo.'],
    }


@pytest.mark.parametrize('value', [float('nan'), float('inf'), float('-inf')])
def test_filter_normalization_rejects_non_finite_floats(value):
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters({'amount': value}, allowed=('amount',))

    assert invalid.value.message_dict == {
        'amount': ['Informe um valor de filtro primitivo.'],
    }


def test_filter_normalization_accepts_exact_date_for_date_filter():
    from reports.filters import normalize_report_filters

    expected = date(2026, 7, 31)
    assert normalize_report_filters(
        {'period_end': expected},
        allowed=('period_end',),
    ) == {'period_end': expected}


def test_filter_normalization_preserves_zero_and_false_and_discards_only_blanks():
    from reports.filters import normalize_report_filters

    result = normalize_report_filters(
        {
            'minimum': 0,
            'active': False,
            'optional': None,
            'empty': '',
        },
        allowed=('minimum', 'active', 'optional', 'empty'),
    )

    assert result == {'minimum': 0, 'active': False}


def test_filter_normalization_reports_required_and_invalid_dates_in_pt_br():
    from reports.filters import normalize_report_filters

    with pytest.raises(ValidationError) as invalid:
        normalize_report_filters(
            {'period_start': '31/07/2026', 'status': ''},
            allowed=('period_start', 'status', 'customer'),
            required=('period_start', 'status', 'customer'),
        )

    assert invalid.value.message_dict == {
        'required_filters': [
            'Filtros obrigatórios ausentes: customer, status.',
        ],
        'period_start': ['Informe uma data válida.'],
    }


CURATED_EXECUTOR_CONTRACTS = {
    'finance.receivables_open_overdue': (
        'Contas a receber em aberto e vencidas',
        ('title_number', 'partner', 'due_date', 'status', 'original_amount', 'open_amount'),
    ),
    'finance.payables_open_overdue': (
        'Contas a pagar em aberto e vencidas',
        ('title_number', 'partner', 'due_date', 'status', 'original_amount', 'open_amount'),
    ),
    'finance.cash_flow': (
        'Fluxo de caixa realizado e projetado',
        ('cash_date', 'flow_type', 'direction', 'account', 'description', 'amount'),
    ),
    'finance.period_result': (
        'Resultado financeiro por período',
        ('period', 'inflow', 'outflow', 'net_result'),
    ),
    'fiscal.documents': (
        'Documentos fiscais por período e situação',
        ('issue_date', 'number', 'series', 'direction', 'partner', 'status', 'total_amount'),
    ),
    'fiscal.tax_assessment': (
        'Apuração de tributos',
        ('period', 'tax_kind', 'debit_amount', 'credit_amount', 'amount_due', 'status'),
    ),
    'fiscal.books': (
        'Livro de entradas e saídas',
        ('entry_date', 'book_type', 'document', 'partner', 'tax_base', 'tax_amount'),
    ),
    'inventory.position': (
        'Posição de estoque',
        (
            'product',
            'lot',
            'expiry_date',
            'warehouse',
            'location',
            'quality_status',
            'quantity',
            'reserved',
            'available',
        ),
    ),
    'inventory.expiry': (
        'Lotes próximos do vencimento ou vencidos',
        ('product', 'lot', 'expiry_date', 'days_to_expiry', 'quality_status', 'quantity'),
    ),
    'inventory.genealogy': (
        'Genealogia e rastreabilidade de lotes',
        (
            'input_product',
            'input_lot',
            'output_product',
            'output_lot',
            'relation_type',
            'quantity',
            'production_order',
        ),
    ),
    'procurement.open_delayed_orders': (
        'Pedidos de compra abertos ou atrasados',
        (
            'order_number',
            'supplier',
            'issue_date',
            'expected_delivery_date',
            'status',
            'total_amount',
            'days_late',
        ),
    ),
    'procurement.receipt_supplier_performance': (
        'Divergências de recebimento e fornecedores',
        (
            'supplier',
            'receipt',
            'product',
            'received',
            'accepted',
            'rejected',
            'acceptance_percent',
        ),
    ),
    'production.orders_status_delay': (
        'Ordens de produção por situação e atraso',
        (
            'order_number',
            'product',
            'priority',
            'responsible',
            'scheduled_end',
            'status',
            'days_late',
        ),
    ),
    'production.consumption_variance': (
        'Consumo planejado versus realizado',
        ('order_number', 'material', 'lot', 'planned', 'actual', 'loss', 'returned', 'variance'),
    ),
    'production.yield_loss_cost': (
        'Rendimento, perdas e custo por ordem',
        (
            'order_number',
            'product',
            'planned',
            'actual_yield',
            'yield_percent',
            'loss',
            'rework',
            'actual_cost',
            'cost_variance',
        ),
    ),
}


@pytest.fixture
def report_context(django_user_model):
    from reports.contracts import ReportContext

    user = django_user_model.objects.create_user(
        username='report.executor@example.com',
        email='report.executor@example.com',
    )
    return ReportContext(filters={}, user=user)


@pytest.fixture
def seeded_report_domains(report_context):
    from auxiliary.models import City, StateProvince
    from costing.models import ProductionCostCapture
    from finance.models import (
        CashFlowEntry,
        ChartOfAccount,
        FinancialAccount,
        FinancialCategory,
        FinancialTitle,
    )
    from fiscal.models import FiscalBookEntry, FiscalCompany, FiscalDocument, TaxAssessmentPeriod
    from formulations.models import ManufacturingRoute, MasterFormula
    from inventory.models import StockBalance, StockLot, StockLotGenealogy, StockQualityStatus
    from masters.models import (
        BusinessPartner,
        Product,
        Site,
        StorageLocation,
        UnitOfMeasure,
        Warehouse,
    )
    from procurement.models import (
        PurchaseOrder,
        PurchaseOrderItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
    )
    from production.models import MaterialConsumption, ProductionOrder, ProductionOutput

    today = date(2026, 7, 28)
    now = timezone.make_aware(datetime(2026, 7, 28, 12, 0))

    def persist(instance):
        instance.full_clean()
        instance.save()
        return instance

    unit = persist(UnitOfMeasure(code='KG-REP', name='Quilograma', symbol='kg'))
    finished = persist(
        Product(
            code='PA-REP',
            description='Produto acabado para relatórios',
            item_type=Product.ItemType.FINISHED_PRODUCT,
            unit=unit,
            status=Product.Status.APPROVED,
        )
    )
    material = persist(
        Product(
            code='MP-REP',
            description='Matéria-prima para relatórios',
            item_type=Product.ItemType.RAW_MATERIAL,
            unit=unit,
            status=Product.Status.APPROVED,
        )
    )
    supplier = persist(
        BusinessPartner(
            code='FOR-REP',
            legal_name='Fornecedor Relatórios Ltda.',
            partner_type=BusinessPartner.PartnerType.SUPPLIER,
            qualification_status=BusinessPartner.QualificationStatus.QUALIFIED,
            qualification_valid_until=today + timedelta(days=365),
        )
    )
    customer = persist(
        BusinessPartner(
            code='CLI-REP',
            legal_name='Cliente Relatórios S.A.',
            partner_type=BusinessPartner.PartnerType.CUSTOMER,
            qualification_status=BusinessPartner.QualificationStatus.QUALIFIED,
            qualification_valid_until=today + timedelta(days=365),
        )
    )
    site = persist(Site(code='PL-REP', name='Planta Relatórios'))
    warehouse = persist(
        Warehouse(
            site=site,
            code='ALM-REP',
            name='Almoxarifado Relatórios',
            warehouse_type=Warehouse.WarehouseType.RAW_MATERIAL,
        )
    )
    location = persist(StorageLocation(warehouse=warehouse, code='A-01', name='Endereço A-01'))
    second_location = persist(
        StorageLocation(warehouse=warehouse, code='A-02', name='Endereço A-02')
    )

    chart = persist(
        ChartOfAccount(
            code='1.1.REP',
            name='Conta de relatórios',
            account_type=ChartOfAccount.AccountType.ASSET,
        )
    )
    category = persist(
        FinancialCategory(
            code='CAT-REP',
            name='Categoria de relatórios',
            category_type=FinancialCategory.CategoryType.BOTH,
            chart_account=chart,
        )
    )
    account = persist(
        FinancialAccount(
            code='BCO-REP',
            name='Banco Relatórios',
            account_type=FinancialAccount.AccountType.BANK,
        )
    )
    receivable = persist(
        FinancialTitle(
            title_number='REC-REP',
            title_type=FinancialTitle.TitleType.RECEIVABLE,
            partner=customer,
            category=category,
            status=FinancialTitle.Status.OVERDUE,
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
            original_amount=Decimal('1000.0000'),
            open_amount=Decimal('1000.0000'),
        )
    )
    payable = persist(
        FinancialTitle(
            title_number='PAG-REP',
            title_type=FinancialTitle.TitleType.PAYABLE,
            partner=supplier,
            category=category,
            status=FinancialTitle.Status.APPROVED,
            issue_date=today - timedelta(days=20),
            due_date=today + timedelta(days=2),
            original_amount=Decimal('300.0000'),
            open_amount=Decimal('300.0000'),
            approved_by=report_context.user,
            approved_at=now,
        )
    )
    persist(
        CashFlowEntry(
            flow_type=CashFlowEntry.FlowType.REALIZED,
            direction=CashFlowEntry.Direction.INFLOW,
            title=receivable,
            financial_account=account,
            cash_date=today,
            amount=Decimal('1000.0000'),
            status=CashFlowEntry.Status.REALIZED,
            description='Recebimento realizado',
        )
    )
    persist(
        CashFlowEntry(
            flow_type=CashFlowEntry.FlowType.REALIZED,
            direction=CashFlowEntry.Direction.OUTFLOW,
            title=payable,
            financial_account=account,
            cash_date=today,
            amount=Decimal('300.0000'),
            status=CashFlowEntry.Status.REALIZED,
            description='Pagamento realizado',
        )
    )

    order = persist(
        PurchaseOrder(
            order_number='PC-REP',
            supplier=supplier,
            status=PurchaseOrder.Status.SENT,
            issue_date=today - timedelta(days=10),
            expected_delivery_date=today - timedelta(days=2),
            total_amount=Decimal('500.0000'),
        )
    )
    order_item = persist(
        PurchaseOrderItem(
            order=order,
            product=material,
            quantity=Decimal('50.0000'),
            unit=unit,
            unit_price=Decimal('10.0000'),
        )
    )
    receipt = persist(
        PurchaseReceipt(
            receipt_number='RC-REP',
            order=order,
            status=PurchaseReceipt.Status.RECEIVED,
            physical_received_at=now,
            received_by=report_context.user,
        )
    )
    receipt_item = persist(
        PurchaseReceiptItem(
            receipt=receipt,
            order_item=order_item,
            product=material,
            received_quantity=Decimal('50.0000'),
            accepted_quantity=Decimal('48.0000'),
            rejected_quantity=Decimal('2.0000'),
            unit=unit,
            lot_number='MP-LOT-REP',
            expiry_date=today + timedelta(days=30),
        )
    )

    state = StateProvince.objects.create(name='Pernambuco Relatórios')
    city = City.objects.create(name='Recife Relatórios', state=state)
    company = persist(
        FiscalCompany(
            legal_name='RGN Farma Relatórios',
            document='12345678000199',
            tax_regime=FiscalCompany.TaxRegime.LUCRO_REAL,
            state_ref=state,
            city_ref=city,
        )
    )
    fiscal_document = persist(
        FiscalDocument(
            company=company,
            partner=supplier,
            document_type=FiscalDocument.DocumentType.INBOUND,
            operation_type=FiscalDocument.OperationType.PURCHASE,
            number='NF-REP',
            series='1',
            issue_date=today,
            operation_date=today,
            status=FiscalDocument.Status.POSTED,
            total_products=Decimal('450.0000'),
            total_taxes=Decimal('50.0000'),
            total_amount=Decimal('500.0000'),
            posted_by=report_context.user,
            posted_at=now,
        )
    )
    persist(
        FiscalBookEntry(
            document=fiscal_document,
            book_type=FiscalBookEntry.BookType.INBOUND,
            entry_date=today,
            total_amount=Decimal('500.0000'),
            tax_amount=Decimal('50.0000'),
        )
    )
    persist(
        TaxAssessmentPeriod(
            period_year=today.year,
            period_month=today.month,
            tax_kind='icms',
            status=TaxAssessmentPeriod.Status.CALCULATED,
            debit_amount=Decimal('180.0000'),
            credit_amount=Decimal('50.0000'),
            balance_amount=Decimal('130.0000'),
            calculated_at=now,
        )
    )

    formula = persist(
        MasterFormula(
            product=finished,
            code='FM-REP',
            version=1,
            status=MasterFormula.Status.APPROVED,
            batch_size=Decimal('100.0000'),
            batch_unit=unit,
            effective_from=today - timedelta(days=30),
            approved_by=report_context.user,
            approved_at=now,
        )
    )
    route = persist(
        ManufacturingRoute(
            product=finished,
            formula=formula,
            code='RT-REP',
            version=1,
            status=ManufacturingRoute.Status.APPROVED,
            effective_from=today - timedelta(days=30),
        )
    )
    production_order = persist(
        ProductionOrder(
            order_number='OP-REP',
            batch_number='PA-LOT-REP',
            product=finished,
            formula=formula,
            route=route,
            planned_quantity=Decimal('100.0000'),
            unit=unit,
            status=ProductionOrder.Status.IN_PROGRESS,
            priority=ProductionOrder.Priority.HIGH,
            scheduled_start=today - timedelta(days=5),
            scheduled_end=today - timedelta(days=1),
            actual_start=now - timedelta(days=2),
            responsible=report_context.user,
            approved_by=report_context.user,
            approved_at=now - timedelta(days=4),
            released_by=report_context.user,
            released_at=now - timedelta(days=3),
            started_by=report_context.user,
            real_loss_quantity=Decimal('5.0000'),
            rework_quantity=Decimal('2.0000'),
        )
    )
    input_lot = persist(
        StockLot(
            product=material,
            lot_number='MP-LOT-REP',
            supplier=supplier,
            quality_status=StockQualityStatus.APPROVED,
            manufacturing_date=today - timedelta(days=300),
            expiry_date=today + timedelta(days=30),
            source_purchase_receipt_item=receipt_item,
        )
    )
    persist(
        StockBalance(
            product=material,
            lot=input_lot,
            warehouse=warehouse,
            location=location,
            quality_status=StockQualityStatus.APPROVED,
            quantity=Decimal('30.0000'),
            reserved_quantity=Decimal('5.0000'),
            unit=unit,
        )
    )
    persist(
        StockBalance(
            product=material,
            lot=input_lot,
            warehouse=warehouse,
            location=second_location,
            quality_status=StockQualityStatus.APPROVED,
            quantity=Decimal('20.0000'),
            reserved_quantity=Decimal('0.0000'),
            unit=unit,
        )
    )
    material_consumption = persist(
        MaterialConsumption(
            order=production_order,
            material=material,
            planned_quantity=Decimal('10.0000'),
            actual_quantity=Decimal('12.0000'),
            loss_quantity=Decimal('1.0000'),
            returned_quantity=Decimal('0.5000'),
            stock_lot=input_lot,
            warehouse=warehouse,
            location=location,
            unit=unit,
            lot_number=input_lot.lot_number,
            quality_status=MaterialConsumption.QualityStatus.APPROVED,
            expiry_date=input_lot.expiry_date,
        )
    )
    persist(
        ProductionOutput(
            order=production_order,
            product=finished,
            lot_number=production_order.batch_number,
            planned_quantity=Decimal('100.0000'),
            produced_quantity=Decimal('90.0000'),
            unit=unit,
            status=ProductionOutput.Status.PENDING,
        )
    )
    production_order.status = ProductionOrder.Status.COMPLETED
    production_order.actual_end = now
    production_order.actual_yield_quantity = Decimal('90.0000')
    production_order.completed_by = report_context.user
    production_order.full_clean()
    production_order.save(
        update_fields=[
            'status',
            'actual_end',
            'actual_yield_quantity',
            'completed_by',
            'updated_at',
        ]
    )
    output_lot = persist(
        StockLot(
            product=finished,
            lot_number=production_order.batch_number,
            quality_status=StockQualityStatus.QUARANTINE,
            source_production_order=production_order,
            manufacturing_date=today,
            expiry_date=today + timedelta(days=365),
        )
    )
    persist(
        StockLotGenealogy(
            input_lot=input_lot,
            output_lot=output_lot,
            relation_type=StockLotGenealogy.RelationType.CONSUMED_IN_PRODUCTION,
            quantity=Decimal('12.0000'),
            unit=unit,
            production_order=production_order,
        )
    )
    cost_capture = ProductionCostCapture(
        production_order=production_order,
        period_start=today.replace(day=1),
        period_end=today,
        planned_cost=Decimal('900.0000'),
        actual_material_cost=Decimal('800.0000'),
        actual_loss_cost=Decimal('100.0000'),
        actual_labor_cost=Decimal('100.0000'),
    )
    cost_capture.calculate_actuals(save=False)
    persist(cost_capture)

    return {
        'today': today,
        'account': account,
        'category': category,
        'receivable': receivable,
        'payable': payable,
        'supplier': supplier,
        'customer': customer,
        'material': material,
        'finished': finished,
        'input_lot': input_lot,
        'production_order': production_order,
        'material_consumption': material_consumption,
        'receipt_item': receipt_item,
    }


@pytest.mark.django_db
@pytest.mark.parametrize(
    ('key', 'expected'),
    tuple(CURATED_EXECUTOR_CONTRACTS.items()),
)
def test_curated_executor_returns_real_rows_with_exact_contract(
    key,
    expected,
    report_context,
    seeded_report_domains,
):
    import reports.executors  # noqa: F401
    from reports.registry import get_executor

    expected_title, expected_keys = expected
    dataset = get_executor(key)(report_context)
    rows = list(dataset.rows)

    assert dataset.title == expected_title
    assert tuple(column.key for column in dataset.columns) == expected_keys
    assert rows
    assert all(tuple(row) == expected_keys for row in rows)


@pytest.mark.django_db
def test_period_result_aggregates_realized_decimal_values(report_context, seeded_report_domains):
    from dataclasses import replace

    from reports.executors.finance import period_result

    today = seeded_report_domains['today']
    context = replace(
        report_context,
        filters={'period_start': today, 'period_end': today},
    )
    row = list(period_result(context).rows)[0]

    assert row['inflow'] == Decimal('1000.0000')
    assert row['outflow'] == Decimal('300.0000')
    assert row['net_result'] == Decimal('700.0000')


@pytest.mark.django_db
def test_executor_decimal_calculations_and_zero_denominators(report_context, seeded_report_domains):
    from dataclasses import replace

    from reports.executors.procurement import _percentage
    from reports.executors.production import consumption_variance, yield_loss_cost

    material_context = replace(
        report_context,
        filters={'product': seeded_report_domains['material'].pk},
    )
    finished_context = replace(
        report_context,
        filters={'product': seeded_report_domains['finished'].pk},
    )
    variance = list(consumption_variance(material_context).rows)[0]
    result = list(yield_loss_cost(finished_context).rows)[0]

    assert variance['variance'] == Decimal('2.0000')
    assert result['actual_yield'] == Decimal('90.0000')
    assert result['yield_percent'] == Decimal('90.0000')
    assert result['actual_cost'] == Decimal('1000.0000')
    assert result['cost_variance'] == Decimal('100.0000')
    assert _percentage(Decimal('1.0000'), Decimal('0.0000')) == Decimal('0.0000')


@pytest.mark.django_db
def test_executor_filters_isolate_dates_status_and_foreign_ids(
    report_context,
    seeded_report_domains,
):
    from dataclasses import replace

    from reports.executors.finance import receivables_open_overdue
    from reports.executors.fiscal import documents
    from reports.executors.inventory import position
    from reports.executors.procurement import open_delayed_orders
    from reports.executors.production import orders_status_delay

    data = seeded_report_domains
    today = data['today']

    assert (
        len(
            list(
                receivables_open_overdue(
                    replace(report_context, filters={'customer': data['customer'].pk})
                ).rows
            )
        )
        == 1
    )
    assert not list(
        receivables_open_overdue(
            replace(report_context, filters={'customer': data['supplier'].pk})
        ).rows
    )
    assert not list(
        documents(
            replace(
                report_context,
                filters={
                    'period_start': today + timedelta(days=1),
                    'status': 'posted',
                    'supplier': data['supplier'].pk,
                },
            )
        ).rows
    )
    assert (
        len(
            list(
                position(
                    replace(
                        report_context,
                        filters={
                            'product': data['material'].pk,
                            'lot': data['input_lot'].pk,
                            'status': 'approved',
                        },
                    )
                ).rows
            )
        )
        == 2
    )
    assert (
        len(
            list(
                open_delayed_orders(
                    replace(
                        report_context,
                        filters={
                            'supplier': data['supplier'].pk,
                            'product': data['material'].pk,
                            'status': 'sent',
                        },
                    )
                ).rows
            )
        )
        == 1
    )
    assert (
        len(
            list(
                orders_status_delay(
                    replace(
                        report_context,
                        filters={
                            'product': data['finished'].pk,
                            'status': 'completed',
                            'period_start': today - timedelta(days=2),
                            'period_end': today,
                        },
                    )
                ).rows
            )
        )
        == 1
    )


@pytest.mark.django_db
def test_finance_executors_enforce_open_types_and_realized_projection_semantics(
    report_context,
    seeded_report_domains,
):
    from dataclasses import replace

    from finance.models import CashFlowEntry, FinancialTitle
    from reports.executors.finance import (
        cash_flow,
        payables_open_overdue,
        period_result,
        receivables_open_overdue,
    )

    data = seeded_report_domains
    today = data['today']
    settled = FinancialTitle(
        title_number='REC-SETTLED-REP',
        title_type=FinancialTitle.TitleType.RECEIVABLE,
        partner=data['customer'],
        category=data['category'],
        status=FinancialTitle.Status.SETTLED,
        issue_date=today - timedelta(days=10),
        due_date=today,
        original_amount=Decimal('50.0000'),
        open_amount=Decimal('0.0000'),
        paid_amount=Decimal('50.0000'),
    )
    settled.full_clean()
    settled.save()
    planned = CashFlowEntry(
        flow_type=CashFlowEntry.FlowType.PLANNED,
        direction=CashFlowEntry.Direction.OUTFLOW,
        title=data['payable'],
        financial_account=data['account'],
        cash_date=today,
        amount=Decimal('200.0000'),
        status=CashFlowEntry.Status.FORECAST,
        description='Projeção válida',
    )
    planned.full_clean()
    planned.save()
    cancelled = CashFlowEntry(
        flow_type=CashFlowEntry.FlowType.REALIZED,
        direction=CashFlowEntry.Direction.INFLOW,
        title=data['receivable'],
        financial_account=data['account'],
        cash_date=today,
        amount=Decimal('999.0000'),
        status=CashFlowEntry.Status.CANCELLED,
        description='Fluxo cancelado',
    )
    cancelled.full_clean()
    cancelled.save()

    customer_context = replace(
        report_context,
        filters={'customer': data['customer'].pk},
    )
    supplier_context = replace(
        report_context,
        filters={'supplier': data['supplier'].pk},
    )
    period_context = replace(
        report_context,
        filters={'period_start': today, 'period_end': today},
    )

    assert [row['title_number'] for row in receivables_open_overdue(customer_context).rows] == [
        'REC-REP'
    ]
    assert [row['title_number'] for row in payables_open_overdue(supplier_context).rows] == [
        'PAG-REP'
    ]
    descriptions = {row['description'] for row in cash_flow(period_context).rows}
    assert 'Projeção válida' in descriptions
    assert 'Fluxo cancelado' not in descriptions
    result = list(period_result(period_context).rows)[0]
    assert result['inflow'] == Decimal('1000.0000')
    assert result['outflow'] == Decimal('300.0000')


@pytest.mark.django_db
def test_expiry_aggregates_only_balances_from_the_reported_lot(
    report_context,
    seeded_report_domains,
):
    from dataclasses import replace

    from reports.executors.inventory import expiry

    data = seeded_report_domains
    rows = list(
        expiry(
            replace(
                report_context,
                filters={'lot': data['input_lot'].pk},
            )
        ).rows
    )

    assert len(rows) == 1
    assert rows[0]['quantity'] == Decimal('50.0000')


@pytest.mark.django_db
def test_day_based_executors_capture_localdate_once(
    report_context,
    seeded_report_domains,
):
    from unittest import mock

    from reports.executors.inventory import expiry
    from reports.executors.procurement import open_delayed_orders
    from reports.executors.production import orders_status_delay

    today = seeded_report_domains['today']
    with mock.patch(
        'reports.executors.inventory.timezone.localdate',
        return_value=today,
    ) as inventory_clock:
        list(expiry(report_context).rows)
    with mock.patch(
        'reports.executors.procurement.timezone.localdate',
        return_value=today,
    ) as procurement_clock:
        list(open_delayed_orders(report_context).rows)
    with mock.patch(
        'reports.executors.production.timezone.localdate',
        return_value=today,
    ) as production_clock:
        list(orders_status_delay(report_context).rows)

    assert inventory_clock.call_count == 1
    assert procurement_clock.call_count == 1
    assert production_clock.call_count == 1


@pytest.mark.django_db
def test_relation_heavy_executors_use_scale_invariant_query_counts(
    report_context,
    seeded_report_domains,
):
    from dataclasses import replace

    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from reports.executors.inventory import position
    from reports.executors.production import yield_loss_cost

    data = seeded_report_domains
    stock_context = replace(
        report_context,
        filters={'product': data['material'].pk},
    )
    production_context = replace(
        report_context,
        filters={'product': data['finished'].pk},
    )

    with CaptureQueriesContext(connection) as stock_queries:
        stock_rows = list(position(stock_context).rows)
    with CaptureQueriesContext(connection) as production_queries:
        production_rows = list(yield_loss_cost(production_context).rows)

    assert len(stock_rows) == 2
    assert len(stock_queries) == 1
    assert len(production_rows) == 1
    assert len(production_queries) == 3


def test_curated_executor_imports_are_static_and_idempotent():
    import reports.executors as executor_package
    from reports.registry import get_executor

    before = {key: get_executor(key) for key in CURATED_EXECUTOR_CONTRACTS}

    reloaded = importlib.reload(executor_package)

    assert reloaded is executor_package
    assert {key: get_executor(key) for key in CURATED_EXECUTOR_CONTRACTS} == before


@pytest.mark.django_db
def test_open_financial_titles_exclude_zero_balance_and_keep_positive_partial_balance(
    report_context,
    seeded_report_domains,
):
    from finance.models import FinancialTitle
    from reports.executors.finance import payables_open_overdue, receivables_open_overdue

    data = seeded_report_domains
    today = data['today']

    def persist_title(*, number, title_type, partner, status, open_amount, paid_amount):
        title = FinancialTitle(
            title_number=number,
            title_type=title_type,
            partner=partner,
            category=data['category'],
            status=status,
            issue_date=today - timedelta(days=10),
            due_date=today,
            original_amount=Decimal('100.0000'),
            open_amount=open_amount,
            paid_amount=paid_amount,
        )
        title.full_clean()
        title.save()
        return title

    persist_title(
        number='REC-ZERO-OPEN',
        title_type=FinancialTitle.TitleType.RECEIVABLE,
        partner=data['customer'],
        status=FinancialTitle.Status.PENDING,
        open_amount=Decimal('0.0000'),
        paid_amount=Decimal('100.0000'),
    )
    persist_title(
        number='PAG-ZERO-OPEN',
        title_type=FinancialTitle.TitleType.PAYABLE,
        partner=data['supplier'],
        status=FinancialTitle.Status.PENDING,
        open_amount=Decimal('0.0000'),
        paid_amount=Decimal('100.0000'),
    )
    persist_title(
        number='REC-PARTIAL-POSITIVE',
        title_type=FinancialTitle.TitleType.RECEIVABLE,
        partner=data['customer'],
        status=FinancialTitle.Status.PARTIALLY_SETTLED,
        open_amount=Decimal('25.0000'),
        paid_amount=Decimal('75.0000'),
    )
    persist_title(
        number='PAG-PARTIAL-POSITIVE',
        title_type=FinancialTitle.TitleType.PAYABLE,
        partner=data['supplier'],
        status=FinancialTitle.Status.PARTIALLY_SETTLED,
        open_amount=Decimal('25.0000'),
        paid_amount=Decimal('75.0000'),
    )

    receivable_numbers = {
        row['title_number'] for row in receivables_open_overdue(report_context).rows
    }
    payable_numbers = {row['title_number'] for row in payables_open_overdue(report_context).rows}

    assert 'REC-ZERO-OPEN' not in receivable_numbers
    assert 'PAG-ZERO-OPEN' not in payable_numbers
    assert 'REC-PARTIAL-POSITIVE' in receivable_numbers
    assert 'PAG-PARTIAL-POSITIVE' in payable_numbers


CURATED_FILTER_VALIDATION_CASES = (
    ('finance.receivables_open_overdue', 'customer', 'foreign_key'),
    ('finance.receivables_open_overdue', 'status', 'choice'),
    ('finance.payables_open_overdue', 'supplier', 'foreign_key'),
    ('finance.payables_open_overdue', 'status', 'choice'),
    ('finance.cash_flow', 'customer', 'foreign_key'),
    ('finance.cash_flow', 'supplier', 'foreign_key'),
    ('finance.cash_flow', 'status', 'choice'),
    ('finance.period_result', 'customer', 'foreign_key'),
    ('finance.period_result', 'supplier', 'foreign_key'),
    ('finance.period_result', 'status', 'choice'),
    ('fiscal.documents', 'supplier', 'foreign_key'),
    ('fiscal.documents', 'customer', 'foreign_key'),
    ('fiscal.documents', 'status', 'choice'),
    ('fiscal.tax_assessment', 'status', 'choice'),
    ('fiscal.books', 'supplier', 'foreign_key'),
    ('fiscal.books', 'customer', 'foreign_key'),
    ('fiscal.books', 'status', 'choice'),
    ('inventory.position', 'product', 'foreign_key'),
    ('inventory.position', 'lot', 'foreign_key'),
    ('inventory.position', 'supplier', 'foreign_key'),
    ('inventory.position', 'status', 'choice'),
    ('inventory.expiry', 'product', 'foreign_key'),
    ('inventory.expiry', 'lot', 'foreign_key'),
    ('inventory.expiry', 'supplier', 'foreign_key'),
    ('inventory.expiry', 'status', 'choice'),
    ('inventory.genealogy', 'product', 'foreign_key'),
    ('inventory.genealogy', 'lot', 'foreign_key'),
    ('procurement.open_delayed_orders', 'supplier', 'foreign_key'),
    ('procurement.open_delayed_orders', 'product', 'foreign_key'),
    ('procurement.open_delayed_orders', 'status', 'choice'),
    ('procurement.receipt_supplier_performance', 'supplier', 'foreign_key'),
    ('procurement.receipt_supplier_performance', 'product', 'foreign_key'),
    ('procurement.receipt_supplier_performance', 'status', 'choice'),
    ('production.orders_status_delay', 'product', 'foreign_key'),
    ('production.orders_status_delay', 'status', 'choice'),
    ('production.consumption_variance', 'product', 'foreign_key'),
    ('production.consumption_variance', 'lot', 'foreign_key'),
    ('production.consumption_variance', 'status', 'choice'),
    ('production.yield_loss_cost', 'product', 'foreign_key'),
    ('production.yield_loss_cost', 'status', 'choice'),
)


@pytest.mark.django_db
@pytest.mark.parametrize(('executor_key', 'field', 'filter_kind'), CURATED_FILTER_VALIDATION_CASES)
@pytest.mark.parametrize(
    'invalid_value',
    (
        0,
        False,
        -1,
        1.5,
        UUID('12345678-1234-5678-1234-567812345678'),
        'invalid-value',
    ),
)
def test_curated_executor_rejects_invalid_controlled_filters_before_query(
    executor_key,
    field,
    filter_kind,
    invalid_value,
    report_context,
):
    from dataclasses import replace

    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    import reports.executors  # noqa: F401
    from reports.registry import get_executor

    context = replace(report_context, filters={field: invalid_value})

    with CaptureQueriesContext(connection) as queries:
        with pytest.raises(ValidationError) as invalid:
            dataset = get_executor(executor_key)(context)
            list(dataset.rows)

    assert field in invalid.value.message_dict
    assert len(queries) == 0


@pytest.mark.django_db
@pytest.mark.parametrize('value_kind', ('integer', 'numeric_string'))
@pytest.mark.parametrize(
    ('executor_key', 'field', 'fixture_key', 'expected_count'),
    (
        ('finance.receivables_open_overdue', 'customer', 'customer', 1),
        ('finance.payables_open_overdue', 'supplier', 'supplier', 1),
        ('inventory.position', 'product', 'material', 2),
        ('inventory.expiry', 'lot', 'input_lot', 1),
    ),
)
def test_relation_filters_accept_positive_integer_and_numeric_string(
    executor_key,
    field,
    fixture_key,
    expected_count,
    value_kind,
    report_context,
    seeded_report_domains,
):
    from dataclasses import replace

    import reports.executors  # noqa: F401
    from reports.registry import get_executor

    identifier = seeded_report_domains[fixture_key].pk
    value = identifier if value_kind == 'integer' else str(identifier)
    context = replace(report_context, filters={field: value})

    assert len(list(get_executor(executor_key)(context).rows)) == expected_count


@pytest.mark.django_db
@pytest.mark.parametrize(
    ('executor_key', 'status'),
    (
        ('finance.receivables_open_overdue', 'overdue'),
        ('finance.payables_open_overdue', 'approved'),
        ('finance.cash_flow', 'realized'),
        ('finance.period_result', 'realized'),
        ('fiscal.documents', 'posted'),
        ('fiscal.tax_assessment', 'calculated'),
        ('fiscal.books', 'posted'),
        ('inventory.position', 'approved'),
        ('inventory.expiry', 'approved'),
        ('procurement.open_delayed_orders', 'sent'),
        ('procurement.receipt_supplier_performance', 'received'),
        ('production.orders_status_delay', 'completed'),
        ('production.consumption_variance', 'completed'),
        ('production.yield_loss_cost', 'completed'),
    ),
)
def test_choice_filters_accept_actual_model_choice(
    executor_key,
    status,
    report_context,
    seeded_report_domains,
):
    from dataclasses import replace

    import reports.executors  # noqa: F401
    from reports.registry import get_executor

    context = replace(report_context, filters={'status': status})

    assert list(get_executor(executor_key)(context).rows)


def test_curated_executors_do_not_use_filter_truthiness():
    from reports.executors import finance, fiscal, inventory, procurement, production

    def is_filters_get(call):
        return (
            isinstance(call, ast.Call)
            and isinstance(call.func, ast.Attribute)
            and isinstance(call.func.value, ast.Name)
            and call.func.value.id == 'filters'
            and call.func.attr == 'get'
        )

    for module in (finance, fiscal, inventory, procurement, production):
        tree = ast.parse(inspect.getsource(module))
        for conditional in (
            node for node in ast.walk(tree) if isinstance(node, (ast.If, ast.IfExp, ast.While))
        ):
            assert not any(is_filters_get(node) for node in ast.walk(conditional.test))
            assert not any(isinstance(node, ast.NamedExpr) for node in ast.walk(conditional.test))


@pytest.mark.django_db
def test_production_delay_uses_local_calendar_date_for_actual_end(
    report_context,
    seeded_report_domains,
):
    from datetime import UTC

    from django.utils import timezone

    from reports.executors.production import orders_status_delay

    order = seeded_report_domains['production_order']
    order.scheduled_end = date(2026, 7, 27)
    order.actual_end = datetime(2026, 7, 28, 1, 0, tzinfo=UTC)
    order.save(update_fields=['scheduled_end', 'actual_end', 'updated_at'])

    with timezone.override('America/Recife'):
        row = next(
            row
            for row in orders_status_delay(report_context).rows
            if row['order_number'] == order.order_number
        )

    assert row['days_late'] == 0


BIG_AUTO_FIELD_MAX = 9_223_372_036_854_775_807


@pytest.mark.django_db
@pytest.mark.parametrize(
    'invalid_value',
    (
        pytest.param(BIG_AUTO_FIELD_MAX + 1, id='max-plus-one-integer'),
        pytest.param(10**5000, id='huge-integer'),
        pytest.param(str(BIG_AUTO_FIELD_MAX + 1), id='max-plus-one-string'),
        pytest.param('1' * 20, id='twenty-digit-string'),
        pytest.param('9' * 5000, id='five-thousand-digit-string'),
        pytest.param('0' * 19 + '1', id='twenty-digit-leading-zero-string'),
    ),
)
def test_positive_integer_filter_rejects_values_outside_big_auto_field_before_query(
    invalid_value,
):
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from reports.executors._filters import positive_integer_filter

    with CaptureQueriesContext(connection) as queries:
        with pytest.raises(ValidationError) as invalid:
            positive_integer_filter({'product': invalid_value}, 'product')

    assert invalid.value.message_dict == {
        'product': ['Informe um identificador inteiro positivo válido.'],
    }
    assert len(queries) == 0


@pytest.mark.parametrize(
    ('value', 'expected'),
    (
        (BIG_AUTO_FIELD_MAX, BIG_AUTO_FIELD_MAX),
        (str(BIG_AUTO_FIELD_MAX), BIG_AUTO_FIELD_MAX),
        ('0001', 1),
    ),
)
def test_positive_integer_filter_accepts_big_auto_field_boundary_and_short_leading_zeros(
    value,
    expected,
):
    from reports.executors._filters import positive_integer_filter

    assert positive_integer_filter({'product': value}, 'product') == expected


@pytest.fixture
def sample_render_dataset():
    from datetime import timezone as datetime_timezone

    from reports.contracts import ReportColumn, ReportDataset

    return ReportDataset(
        title='Relatório farmacêutico',
        columns=(
            ReportColumn('title', 'Título'),
            ReportColumn('date', 'Data', 'date'),
            ReportColumn('timestamp', 'Momento', 'datetime'),
            ReportColumn('amount', 'Valor', 'decimal'),
            ReportColumn('count', 'Quantidade', 'integer'),
            ReportColumn('active', 'Ativo'),
            ReportColumn('optional', 'Opcional'),
        ),
        rows=(
            {
                'title': 'Cápsula ação prolongada',
                'date': date(2026, 7, 28),
                'timestamp': datetime(
                    2026,
                    7,
                    28,
                    12,
                    30,
                    45,
                    tzinfo=datetime_timezone(timedelta(hours=-3)),
                ),
                'amount': Decimal('12.5'),
                'count': 7,
                'active': True,
                'optional': None,
            },
        ),
    )


@pytest.mark.parametrize(
    ('export_format', 'signature', 'mime_type'),
    (
        ('csv', b'\xef\xbb\xbf', 'text/csv; charset=utf-8'),
        (
            'xlsx',
            b'PK',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ),
        ('pdf', b'%PDF', 'application/pdf'),
    ),
)
def test_render_report_creates_valid_formats_with_exact_metadata(
    sample_render_dataset,
    export_format,
    signature,
    mime_type,
):
    from io import BytesIO

    from openpyxl import load_workbook
    from pypdf import PdfReader

    from reports.renderers import render_report

    rendered = render_report(sample_render_dataset, export_format)

    assert rendered.content.startswith(signature)
    assert rendered.mime_type == mime_type
    assert rendered.extension == export_format

    if export_format == 'xlsx':
        workbook = load_workbook(BytesIO(rendered.content), read_only=True)
        try:
            assert workbook.sheetnames == ['Relatório']
            assert workbook.active['A1'].value == 'Título'
        finally:
            workbook.close()
    if export_format == 'pdf':
        assert len(PdfReader(BytesIO(rendered.content)).pages) >= 1


def test_rendered_report_is_frozen_typed_and_slotted(sample_render_dataset):
    from reports.renderers import RenderedReport, render_report

    rendered = render_report(sample_render_dataset, 'csv')

    assert get_type_hints(RenderedReport) == {
        'content': bytes,
        'mime_type': str,
        'extension': str,
    }
    assert not hasattr(rendered, '__dict__')
    with pytest.raises(FrozenInstanceError):
        rendered.extension = 'txt'


@pytest.mark.parametrize(
    'invalid_format',
    (
        '',
        'CSV',
        ' csv',
        0,
        False,
        type('ExportFormatSubclass', (str,), {})('csv'),
    ),
)
def test_render_report_rejects_noncanonical_or_nonexact_format_without_consuming_rows(
    invalid_format,
):
    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    consumed = False

    def rows():
        nonlocal consumed
        consumed = True
        yield {'value': 'não deve ser lido'}

    dataset = ReportDataset(
        title='Inválido',
        columns=(ReportColumn('value', 'Valor'),),
        rows=rows(),
    )

    with pytest.raises(ValidationError) as invalid:
        render_report(dataset, invalid_format)

    assert invalid.value.message_dict == {
        'export_format': ['Formato de exportação não suportado.'],
    }
    assert consumed is False


def test_render_report_does_not_execute_format_dunders(sample_render_dataset):
    from reports.renderers import render_report

    class HostileFormat:
        def __eq__(self, other):
            raise AssertionError('comparison dunder executed')

        def __hash__(self):
            raise AssertionError('hash dunder executed')

        def __str__(self):
            raise AssertionError('string dunder executed')

    with pytest.raises(ValidationError) as invalid:
        render_report(sample_render_dataset, HostileFormat())

    assert invalid.value.message_dict == {
        'export_format': ['Formato de exportação não suportado.'],
    }


@pytest.mark.parametrize('export_format', ('csv', 'xlsx', 'pdf'))
def test_render_report_rejects_missing_columns_without_consuming_rows(export_format):
    from reports.contracts import ReportDataset
    from reports.renderers import render_report

    consumed = False

    def rows():
        nonlocal consumed
        consumed = True
        yield {}

    dataset = ReportDataset(title='Sem colunas', columns=(), rows=rows())

    with pytest.raises(ValidationError) as invalid:
        render_report(dataset, export_format)

    assert invalid.value.message_dict == {
        'columns': ['O relatório deve possuir ao menos uma coluna.'],
    }
    assert consumed is False


@pytest.mark.parametrize('export_format', ('csv', 'xlsx', 'pdf'))
def test_render_report_iterates_rows_exactly_once(export_format):
    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    class CountingRows:
        def __init__(self):
            self.iterations = 0

        def __iter__(self):
            self.iterations += 1
            yield {'value': 'única passagem'}

    rows = CountingRows()
    dataset = ReportDataset(
        title='Passagem única',
        columns=(ReportColumn('value', 'Valor'),),
        rows=rows,
    )

    render_report(dataset, export_format)

    assert rows.iterations == 1


@pytest.mark.parametrize('export_format', ('csv', 'xlsx', 'pdf'))
def test_render_report_accepts_header_only_dataset(export_format):
    from io import BytesIO, StringIO
    import csv

    from openpyxl import load_workbook
    from pypdf import PdfReader

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    dataset = ReportDataset(
        title='Somente cabeçalho',
        columns=(ReportColumn('value', 'Valor'),),
        rows=(),
    )

    rendered = render_report(dataset, export_format)

    if export_format == 'csv':
        parsed = list(
            csv.reader(
                StringIO(rendered.content.decode('utf-8-sig'), newline=''),
                delimiter=';',
            )
        )
        assert parsed == [['Valor']]
    elif export_format == 'xlsx':
        workbook = load_workbook(BytesIO(rendered.content), read_only=True)
        try:
            assert list(workbook.active.values) == [('Valor',)]
        finally:
            workbook.close()
    else:
        assert len(PdfReader(BytesIO(rendered.content)).pages) == 1


def test_csv_uses_bom_semicolon_newline_handling_and_deterministic_cells(
    sample_render_dataset,
):
    import csv
    from io import StringIO

    from reports.renderers import render_report

    rendered = render_report(sample_render_dataset, 'csv')
    rows = list(
        csv.reader(
            StringIO(rendered.content.decode('utf-8-sig'), newline=''),
            delimiter=';',
        )
    )

    assert rows == [
        ['Título', 'Data', 'Momento', 'Valor', 'Quantidade', 'Ativo', 'Opcional'],
        [
            'Cápsula ação prolongada',
            '2026-07-28',
            '2026-07-28T12:30:45-03:00',
            '12.5000',
            '7',
            'True',
            '',
        ],
    ]


@pytest.mark.parametrize(
    'dangerous_text',
    (
        '=SUM(1;1)',
        '+1+1',
        '-2+3',
        '@comando',
        ' =SUM(1;1)',
        '\t=SUM(1;1)',
        '\r+1+1',
        '\n-2+3',
        '\x00@comando',
        '\x01 \t=SUM(1;1)',
    ),
)
@pytest.mark.parametrize('export_format', ('csv', 'xlsx'))
def test_spreadsheet_renderers_neutralize_formula_text_after_leading_whitespace(
    dangerous_text,
    export_format,
):
    import csv
    from io import BytesIO, StringIO

    from openpyxl import load_workbook

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    dataset = ReportDataset(
        title='Fórmulas',
        columns=(ReportColumn('value', 'Valor'),),
        rows=({'value': dangerous_text},),
    )
    rendered = render_report(dataset, export_format)

    if export_format == 'csv':
        value = list(
            csv.reader(
                StringIO(rendered.content.decode('utf-8-sig'), newline=''),
                delimiter=';',
            )
        )[1][0]
        assert value.startswith("'")
    else:
        workbook = load_workbook(BytesIO(rendered.content), read_only=True, data_only=False)
        try:
            cell = workbook.active['A2']
            assert cell.value.startswith("'")
            assert cell.data_type == 's'
        finally:
            workbook.close()


@pytest.mark.parametrize('export_format', ('csv', 'xlsx'))
def test_spreadsheet_renderers_do_not_prefix_negative_numeric_values(export_format):
    import csv
    from io import BytesIO, StringIO

    from openpyxl import load_workbook

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    dataset = ReportDataset(
        title='Negativos',
        columns=(
            ReportColumn('decimal', 'Decimal', 'decimal'),
            ReportColumn('integer', 'Inteiro', 'integer'),
        ),
        rows=({'decimal': Decimal('-1.25'), 'integer': -2},),
    )
    rendered = render_report(dataset, export_format)

    if export_format == 'csv':
        values = list(
            csv.reader(
                StringIO(rendered.content.decode('utf-8-sig'), newline=''),
                delimiter=';',
            )
        )[1]
    else:
        workbook = load_workbook(BytesIO(rendered.content), read_only=True)
        try:
            values = list(next(workbook.active.iter_rows(min_row=2, values_only=True)))
        finally:
            workbook.close()

    assert values == ['-1.2500', '-2']


@pytest.mark.parametrize('export_format', ('csv', 'xlsx'))
def test_spreadsheet_renderers_sanitize_illegal_controls_and_preserve_unicode(
    export_format,
):
    import csv
    from io import BytesIO, StringIO

    from openpyxl import load_workbook

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    dataset = ReportDataset(
        title='Controles',
        columns=(ReportColumn('value', 'Valor'),),
        rows=({'value': 'A\x00B\x01C\x0bD\x1fE\tF\nG\rH ação'},),
    )
    rendered = render_report(dataset, export_format)

    if export_format == 'csv':
        value = list(
            csv.reader(
                StringIO(rendered.content.decode('utf-8-sig'), newline=''),
                delimiter=';',
            )
        )[1][0]
    else:
        workbook = load_workbook(BytesIO(rendered.content), read_only=True)
        try:
            value = workbook.active['A2'].value
        finally:
            workbook.close()

    assert value == 'ABCDE\tF\nG\nH ação'


@pytest.mark.parametrize('export_format', ('csv', 'xlsx', 'pdf'))
def test_render_report_rejects_unsupported_cell_without_executing_string_dunder(
    export_format,
):
    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    class HostileCell:
        def __str__(self):
            raise AssertionError('cell string dunder executed')

    dataset = ReportDataset(
        title='Célula hostil',
        columns=(ReportColumn('value', 'Valor'),),
        rows=({'value': HostileCell()},),
    )

    with pytest.raises(ValidationError) as invalid:
        render_report(dataset, export_format)

    assert invalid.value.message_dict == {
        'rows': ['O relatório contém um valor de célula não suportado.'],
    }


def test_pdf_escapes_title_and_cell_markup_as_literal_text():
    from io import BytesIO

    from pypdf import PdfReader

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    dataset = ReportDataset(
        title='<b>Título & controle</b>',
        columns=(ReportColumn('value', '<i>Cabeçalho</i>'),),
        rows=({'value': '<b>texto</b> & <script>ação</script>'},),
    )

    rendered = render_report(dataset, 'pdf')
    extracted = '\n'.join(
        page.extract_text() or '' for page in PdfReader(BytesIO(rendered.content)).pages
    )

    assert '<b>Título & controle</b>' in extracted
    assert '<i>Cabeçalho</i>' in extracted
    assert '<b>texto</b> & <script>ação</script>' in extracted


def test_pdf_repeats_header_on_multipage_table():
    from io import BytesIO

    from pypdf import PdfReader

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    header = 'Cabeçalho repetido'
    dataset = ReportDataset(
        title='Múltiplas páginas',
        columns=(ReportColumn('value', header),),
        rows=({'value': f'Linha {index:03d} conteúdo'} for index in range(240)),
    )

    rendered = render_report(dataset, 'pdf')
    reader = PdfReader(BytesIO(rendered.content))

    assert len(reader.pages) > 1
    assert all(header in (page.extract_text() or '') for page in reader.pages)


def test_pdf_handles_many_columns_and_long_unbroken_text_without_layout_error():
    from io import BytesIO

    from pypdf import PdfReader

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    columns = tuple(ReportColumn(f'column_{index}', f'Coluna {index}') for index in range(12))
    long_value = 'ação' * 150
    dataset = ReportDataset(
        title='Relatório com título ' + ('muito longo ' * 30),
        columns=columns,
        rows=(
            {column.key: f'{long_value}-{row_index}' for column in columns}
            for row_index in range(3)
        ),
    )

    rendered = render_report(dataset, 'pdf')

    assert len(PdfReader(BytesIO(rendered.content)).pages) >= 1


XLSX_TEXT_LIMIT = 32_767


def test_xlsx_accepts_plain_text_at_excel_cell_limit():
    from io import BytesIO

    from openpyxl import load_workbook

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    value = 'x' * XLSX_TEXT_LIMIT
    dataset = ReportDataset(
        title='Limite XLSX',
        columns=(ReportColumn('value', 'Valor'),),
        rows=({'value': value},),
    )

    rendered = render_report(dataset, 'xlsx')
    workbook = load_workbook(BytesIO(rendered.content), read_only=True)
    try:
        assert workbook.active['A2'].value == value
        assert len(workbook.active['A2'].value) == XLSX_TEXT_LIMIT
    finally:
        workbook.close()


@pytest.mark.parametrize(
    'value',
    (
        pytest.param('x' * 40_000, id='forty-thousand-plain-characters'),
        pytest.param(
            '=' + ('x' * (XLSX_TEXT_LIMIT - 1)),
            id='formula-at-limit-before-apostrophe',
        ),
    ),
)
def test_xlsx_rejects_final_text_above_excel_limit_without_partial_content(
    value,
    monkeypatch,
):
    from io import BytesIO

    from reports.contracts import ReportColumn, ReportDataset
    import reports.renderers as renderers

    streams = []

    def tracked_stream():
        stream = BytesIO()
        streams.append(stream)
        return stream

    monkeypatch.setattr(renderers, 'BytesIO', tracked_stream)
    dataset = ReportDataset(
        title='Acima do limite XLSX',
        columns=(ReportColumn('value', 'Valor'),),
        rows=({'value': value},),
    )

    with pytest.raises(ValidationError) as invalid:
        renderers.render_report(dataset, 'xlsx')

    assert invalid.value.message_dict == {
        'content': ['O relatório contém uma célula que excede o limite do XLSX.'],
    }
    assert len(streams) == 1
    assert streams[0].getvalue() == b''


def test_csv_does_not_apply_xlsx_cell_length_limit():
    import csv
    from io import StringIO

    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    value = 'x' * 40_000
    dataset = ReportDataset(
        title='Texto longo CSV',
        columns=(ReportColumn('value', 'Valor'),),
        rows=({'value': value},),
    )

    rendered = render_report(dataset, 'csv')
    parsed = list(
        csv.reader(
            StringIO(rendered.content.decode('utf-8-sig'), newline=''),
            delimiter=';',
        )
    )

    assert parsed[1][0] == value


@pytest.mark.parametrize('export_format', ('csv', 'xlsx', 'pdf'))
def test_render_report_rejects_hostile_column_key_without_hashing_or_consuming_rows(
    export_format,
):
    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    consumed = False

    class HostileKey(str):
        def __hash__(self):
            raise AssertionError('column key hash dunder executed')

    def rows():
        nonlocal consumed
        consumed = True
        yield {'value': 'não deve ser lido'}

    dataset = ReportDataset(
        title='Key hostil',
        columns=(ReportColumn(HostileKey('value'), 'Valor'),),
        rows=rows(),
    )

    with pytest.raises(ValidationError) as invalid:
        render_report(dataset, export_format)

    assert invalid.value.message_dict == {
        'columns': ['O relatório contém uma coluna inválida.'],
    }
    assert consumed is False


@pytest.mark.parametrize(
    'columns',
    (
        pytest.param(
            (('value', 'Valor'), ('value', 'Valor repetido')),
            id='duplicate-key',
        ),
        pytest.param((('', 'Valor'),), id='empty-key'),
    ),
)
def test_render_report_rejects_duplicate_or_empty_column_key_before_rows(columns):
    from reports.contracts import ReportColumn, ReportDataset
    from reports.renderers import render_report

    consumed = False

    def rows():
        nonlocal consumed
        consumed = True
        yield {'value': 'não deve ser lido'}

    dataset = ReportDataset(
        title='Keys inválidas',
        columns=tuple(ReportColumn(key, label) for key, label in columns),
        rows=rows(),
    )

    with pytest.raises(ValidationError) as invalid:
        render_report(dataset, 'csv')

    assert invalid.value.message_dict == {
        'columns': ['O relatório contém uma coluna inválida.'],
    }
    assert consumed is False


@pytest.fixture
def system_report_definition(db, isolated_report_registry, django_user_model):
    from reports.contracts import ReportColumn, ReportDataset
    from reports.models import ReportDefinition, ReportExecution

    registry, _executors = isolated_report_registry

    @registry.register_executor('tests.lifecycle')
    def lifecycle_executor(context):
        return ReportDataset(
            title='Ciclo de execução',
            columns=(ReportColumn('value', 'Valor'),),
            rows=({'value': 'linha 1'}, {'value': 'linha 2'}),
        )

    owner = django_user_model.objects.create_user(
        username='system.definition@example.com',
        email='system.definition@example.com',
    )
    return ReportDefinition.objects.create(
        code='SYS-LIFECYCLE',
        title='Relatório gerenciado',
        module=ReportDefinition.Module.FINANCE,
        category=ReportDefinition.Category.MANAGEMENT,
        allowed_export_formats=[
            ReportExecution.ExportFormat.CSV,
            ReportExecution.ExportFormat.XLSX,
            ReportExecution.ExportFormat.PDF,
        ],
        query_config={'source': 'finance.FinancialTitle'},
        executor_key='tests.lifecycle',
        filter_schema={'status': {'type': 'text', 'label': 'Situação'}},
        required_permission='finance.view_financialtitle',
        is_system_managed=True,
        owner=owner,
    )


@pytest.mark.django_db
def test_report_execution_engine_schema_and_procurement_choices():
    from django.db.models import PROTECT

    from reports.models import (
        DashboardWidget,
        DashboardWorkspace,
        ReportDefinition,
        ReportExecution,
    )

    assert ReportDefinition.Module.PROCUREMENT == 'procurement'
    assert ('procurement', 'Compras') in ReportDefinition.Module.choices
    for model in (ReportDefinition, DashboardWorkspace, DashboardWidget):
        assert ('procurement', 'Compras') in model._meta.get_field('module').choices

    expected_definition_fields = {
        'executor_key',
        'is_system_managed',
        'filter_schema',
        'required_permission',
    }
    assert expected_definition_fields <= {
        field.name for field in ReportDefinition._meta.concrete_fields
    }
    result_file = ReportExecution._meta.get_field('result_file')
    assert result_file.remote_field.on_delete is PROTECT
    assert result_file.remote_field.related_name == 'report_executions'
    assert result_file.null is True
    assert result_file.blank is True


def test_report_execution_engine_migration_is_deterministic_and_scoped():
    from django.db import migrations, models

    migration = importlib.import_module('reports.migrations.0004_report_execution_engine').Migration
    operations = migration.operations

    assert (
        'reports',
        '0003_remove_dashboardworkspace_unique_tenant_dashboard_workspace_code_and_more',
    ) in migration.dependencies
    assert (
        'files',
        '0004_remove_protectedfile_unique_tenant_protected_file_number_and_more',
    ) in migration.dependencies
    assert [
        operation.name for operation in operations if isinstance(operation, migrations.AddField)
    ] == [
        'executor_key',
        'filter_schema',
        'is_system_managed',
        'required_permission',
        'result_file',
    ]
    assert {
        (operation.model_name, operation.name)
        for operation in operations
        if isinstance(operation, migrations.AlterField)
    } == {
        ('dashboardwidget', 'module'),
        ('dashboardworkspace', 'module'),
        ('reportdefinition', 'module'),
    }
    assert all(
        not isinstance(operation, (migrations.AlterModelOptions, migrations.RunPython))
        for operation in operations
    )
    result_operation = next(
        operation
        for operation in operations
        if isinstance(operation, migrations.AddField) and operation.model_name == 'reportexecution'
    )
    assert isinstance(result_operation.field, models.ForeignKey)
    assert result_operation.field.remote_field.on_delete is models.PROTECT


@pytest.mark.django_db
def test_system_definition_creation_requires_registered_executor_without_registry_leak(
    isolated_report_registry,
):
    from reports.models import ReportDefinition, ReportExecution

    definition = ReportDefinition(
        code='SYS-UNKNOWN',
        title='Executor desconhecido',
        module=ReportDefinition.Module.FINANCE,
        category=ReportDefinition.Category.MANAGEMENT,
        allowed_export_formats=[ReportExecution.ExportFormat.CSV],
        query_config={},
        executor_key='tests.unknown',
        filter_schema={},
        required_permission='finance.view_financialtitle',
        is_system_managed=True,
    )

    with pytest.raises(ValidationError) as invalid:
        definition.full_clean()

    assert invalid.value.message_dict == {
        'executor_key': ['Executor de relatório não registrado.'],
    }


@pytest.mark.parametrize(
    ('field_name', 'replacement'),
    (
        ('code', 'SYS-CHANGED'),
        ('executor_key', 'tests.other'),
        ('query_config', {'source': 'fiscal.FiscalDocument'}),
        ('filter_schema', {'type': 'array'}),
        ('required_permission', 'fiscal.view_fiscaldocument'),
        ('is_system_managed', False),
        ('module', 'fiscal'),
        ('allowed_export_formats', ['pdf']),
    ),
)
@pytest.mark.django_db
def test_system_definition_technical_fields_are_immutable_in_model_clean_and_save(
    system_report_definition,
    field_name,
    replacement,
):
    definition = system_report_definition
    setattr(definition, field_name, replacement)

    with pytest.raises(ValidationError) as invalid_clean:
        definition.full_clean()
    assert field_name in invalid_clean.value.message_dict

    with pytest.raises(ValidationError) as invalid_save:
        definition.save()
    assert field_name in invalid_save.value.message_dict


@pytest.mark.django_db
def test_system_definition_rejects_hostile_technical_json_without_invoking_dunders(
    system_report_definition,
):
    class HostileDict(dict):
        def __bool__(self):
            raise AssertionError('dunder técnico não pode ser chamado')

        def __eq__(self, other):
            raise AssertionError('dunder técnico não pode ser chamado')

        def __iter__(self):
            raise AssertionError('dunder técnico não pode ser chamado')

        def items(self):
            raise AssertionError('dunder técnico não pode ser chamado')

    system_report_definition.query_config = HostileDict()

    with pytest.raises(ValidationError) as invalid:
        system_report_definition.full_clean()

    assert invalid.value.message_dict == {
        'query_config': ['A configuração técnica deve ser um objeto JSON seguro.'],
    }


@pytest.mark.django_db
def test_user_managed_definition_keeps_editable_legacy_contract(django_user_model):
    from reports.models import ReportDefinition, ReportExecution
    from reports.serializers import ReportDefinitionSerializer

    owner = django_user_model.objects.create_user(
        username='user.definition@example.com',
        email='user.definition@example.com',
    )
    definition = ReportDefinition.objects.create(
        code='USR-LEGACY',
        title='Relatório do usuário',
        module=ReportDefinition.Module.QUALITY,
        category=ReportDefinition.Category.OPERATIONAL,
        allowed_export_formats=[ReportExecution.ExportFormat.CSV],
        query_config={'source': 'quality.QualitySample'},
        owner=owner,
    )
    serializer = ReportDefinitionSerializer(
        definition,
        data={
            'code': 'USR-EDITED',
            'query_config': {'source': 'quality.QualityResult'},
        },
        partial=True,
    )

    assert serializer.is_valid(), serializer.errors
    updated = serializer.save()
    assert updated.code == 'USR-EDITED'
    assert updated.query_config == {'source': 'quality.QualityResult'}


@pytest.mark.parametrize(
    'field_name',
    (
        'code',
        'executor_key',
        'query_config',
        'filter_schema',
        'required_permission',
        'is_system_managed',
    ),
)
@pytest.mark.django_db
def test_system_definition_serializer_explicitly_rejects_technical_input(
    system_report_definition,
    field_name,
):
    from reports.serializers import ReportDefinitionSerializer

    serializer = ReportDefinitionSerializer(
        system_report_definition,
        data={field_name: 'attempted-change'},
        partial=True,
    )

    assert serializer.is_valid() is False
    assert serializer.errors[field_name] == [
        'Este campo técnico não pode ser alterado em relatório gerenciado pelo sistema.'
    ]


@pytest.mark.django_db
def test_report_definition_serializer_marks_new_technical_fields_read_only(
    system_report_definition,
):
    from reports.serializers import ReportDefinitionSerializer

    serializer = ReportDefinitionSerializer(system_report_definition)

    for field_name in (
        'executor_key',
        'filter_schema',
        'required_permission',
        'is_system_managed',
    ):
        assert serializer.fields[field_name].read_only is True


@pytest.mark.django_db
def test_system_definition_serializer_can_update_nontechnical_field_without_clone_drift(
    system_report_definition,
):
    from reports.serializers import ReportDefinitionSerializer

    serializer = ReportDefinitionSerializer(
        system_report_definition,
        data={'title': 'Título operacional atualizado'},
        partial=True,
    )

    assert serializer.is_valid(), serializer.errors
    updated = serializer.save()
    assert updated.title == 'Título operacional atualizado'


@pytest.fixture
def report_execution_actor(db, django_user_model):
    from django.contrib.auth.models import Permission

    user = django_user_model.objects.create_user(
        username='report.execution@example.com',
        email='report.execution@example.com',
    )
    permissions = Permission.objects.filter(
        content_type__app_label__in=('reports', 'finance'),
        codename__in=('add_reportexecution', 'view_financialtitle'),
    )
    assert permissions.count() == 2
    user.user_permissions.add(*permissions)
    return user


def _configure_report_encryption(settings, tmp_path):
    import base64

    settings.MEDIA_ROOT = tmp_path
    settings.DATA_ENCRYPTION_KEYS = f'test:{base64.urlsafe_b64encode(b"4" * 32).decode("ascii")}'
    settings.DATA_ENCRYPTION_KEY_ID = 'test'


@pytest.mark.parametrize(
    ('export_format', 'signature', 'mime_type'),
    (
        ('csv', b'\xef\xbb\xbf', 'text/csv; charset=utf-8'),
        (
            'xlsx',
            b'PK',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ),
        ('pdf', b'%PDF', 'application/pdf'),
    ),
)
@pytest.mark.django_db
def test_execution_stores_real_encrypted_protected_artifact_for_every_format(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    export_format,
    signature,
    mime_type,
):
    import hashlib

    from files.models import ProtectedFile, ProtectedFileAuditTrail
    from governance.models import GovernanceAuditLog
    from reports.models import ReportNotification
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format=export_format,
        requested_by=report_execution_actor,
    )

    completed = execute_report(execution, report_execution_actor)
    completed.refresh_from_db()
    protected_file = completed.result_file
    decrypted = protected_file.read_encrypted_content(report_execution_actor)

    assert completed.status == completed.Status.COMPLETED
    assert completed.started_at is not None
    assert completed.completed_at is not None
    assert completed.completed_at >= completed.started_at
    assert completed.error_message == ''
    assert completed.row_count == 2
    assert completed.result_reference == protected_file.file_reference
    assert completed.content_hash == protected_file.content_hash
    assert completed.content_hash == f'sha256:{hashlib.sha256(decrypted).hexdigest()}'
    assert protected_file.file_type == ProtectedFile.FileType.REPORT
    assert protected_file.source_module == ProtectedFile.SourceModule.OPERATIONAL
    assert protected_file.source_model == 'reports.ReportExecution'
    assert protected_file.source_record_id == str(completed.pk)
    assert protected_file.origin == ProtectedFile.Origin.SYSTEM
    assert protected_file.criticality == ProtectedFile.Criticality.MEDIUM
    assert protected_file.confidentiality == ProtectedFile.Confidentiality.INTERNAL
    assert protected_file.responsible == report_execution_actor
    assert protected_file.uploaded_by == report_execution_actor
    assert protected_file.file_name == f'{completed.execution_number}.{export_format}'
    assert protected_file.mime_type == mime_type
    assert protected_file.file_size == len(decrypted)
    assert protected_file.encrypted_size > protected_file.file_size
    assert protected_file.is_encrypted is True
    assert decrypted.startswith(signature)
    encrypted_payload = (tmp_path / protected_file.file_reference).read_bytes()
    assert not encrypted_payload.startswith(signature)
    assert ProtectedFileAuditTrail.objects.filter(
        protected_file=protected_file,
        action=ProtectedFileAuditTrail.Action.UPLOAD,
        actor=report_execution_actor,
    ).exists()
    assert set(
        GovernanceAuditLog.objects.filter(
            module='reports',
            target_model='ReportExecution',
            target_record_id=str(completed.pk),
        ).values_list('action', flat=True)
    ) == {'report.execution.claimed', 'report.execution.completed'}
    assert ReportNotification.objects.filter(
        execution=completed,
        recipient=report_execution_actor,
        status=ReportNotification.Status.SENT,
    ).exists()


@pytest.mark.parametrize(
    ('grant_report_permission', 'grant_domain_permission'),
    ((False, True), (True, False)),
)
@pytest.mark.django_db
def test_execution_requires_report_and_domain_permissions_without_side_effects(
    system_report_definition,
    django_user_model,
    settings,
    tmp_path,
    grant_report_permission,
    grant_domain_permission,
):
    from django.contrib.auth.models import Permission
    from django.core.exceptions import PermissionDenied

    from files.models import ProtectedFile
    from governance.models import GovernanceAuditLog
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    email = f'permission.{grant_report_permission}.{grant_domain_permission}@example.com'
    user = django_user_model.objects.create_user(username=email, email=email)
    codenames = []
    if grant_report_permission:
        codenames.append('add_reportexecution')
    if grant_domain_permission:
        codenames.append('view_financialtitle')
    user.user_permissions.add(*Permission.objects.filter(codename__in=codenames))
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=user,
    )

    with pytest.raises(PermissionDenied):
        execute_report(execution, user)

    execution.refresh_from_db()
    assert execution.status == execution.Status.PENDING
    assert execution.started_at is None
    assert execution.result_file_id is None
    assert (
        ProtectedFile.objects.filter(
            source_model='reports.ReportExecution',
            source_record_id=str(execution.pk),
        ).exists()
        is False
    )
    assert (
        GovernanceAuditLog.objects.filter(
            module='reports',
            target_model='ReportExecution',
            target_record_id=str(execution.pk),
        ).exists()
        is False
    )
    assert list(tmp_path.rglob('*.enc')) == []


@pytest.mark.django_db
def test_execution_rejects_inactive_persisted_actor_even_with_stale_active_instance(
    system_report_definition,
    report_execution_actor,
):
    from django.core.exceptions import PermissionDenied

    from reports.services import execute_report

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    type(report_execution_actor).objects.filter(pk=report_execution_actor.pk).update(
        is_active=False
    )
    assert report_execution_actor.is_active is True

    with pytest.raises(PermissionDenied):
        execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    assert execution.status == execution.Status.PENDING
    assert execution.result_file_id is None


@pytest.mark.django_db
def test_execution_rejects_actor_different_from_persisted_requester(
    system_report_definition,
    report_execution_actor,
    django_user_model,
):
    from django.contrib.auth.models import Permission
    from django.core.exceptions import PermissionDenied

    from reports.services import execute_report

    other = django_user_model.objects.create_user(
        username='other.actor@example.com',
        email='other.actor@example.com',
    )
    other.user_permissions.add(
        *Permission.objects.filter(codename__in=('add_reportexecution', 'view_financialtitle'))
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    with pytest.raises(PermissionDenied):
        execute_report(execution, other)

    execution.refresh_from_db()
    assert execution.status == execution.Status.PENDING
    assert execution.result_file_id is None


@pytest.mark.parametrize('status', ('running', 'completed', 'failed', 'cancelled'))
@pytest.mark.django_db
def test_execution_claim_accepts_only_pending_state(
    system_report_definition,
    report_execution_actor,
    status,
):
    from reports.models import ReportExecution
    from reports.services import execute_report

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    ReportExecution.objects.filter(pk=execution.pk).update(status=status)

    with pytest.raises(ValidationError) as invalid:
        execute_report(execution, report_execution_actor)

    assert invalid.value.message_dict == {
        'status': ['A execução não pode ser processada neste estado.'],
    }
    execution.refresh_from_db()
    assert execution.status == status
    assert execution.result_file_id is None


@pytest.mark.django_db(transaction=True)
def test_concurrent_duplicate_executes_renderer_and_storage_exactly_once(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
    monkeypatch,
):
    from threading import Event, Lock

    from django.db import close_old_connections

    from files.models import ProtectedFile
    from reports.contracts import ReportColumn, ReportDataset
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    _registry, executors = isolated_report_registry
    entered = Event()
    release = Event()
    counter_lock = Lock()
    counts = {'executor': 0, 'store': 0}

    def blocking_executor(context):
        with counter_lock:
            counts['executor'] += 1
        entered.set()
        assert release.wait(timeout=10)
        return ReportDataset(
            title='Concorrência',
            columns=(ReportColumn('value', 'Valor'),),
            rows=({'value': 'única'},),
        )

    executors['tests.lifecycle'] = blocking_executor
    original_store = ProtectedFile.store_encrypted_content

    def counted_store(self, *args, **kwargs):
        with counter_lock:
            counts['store'] += 1
        return original_store(self, *args, **kwargs)

    monkeypatch.setattr(ProtectedFile, 'store_encrypted_content', counted_store)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    def first_attempt():
        close_old_connections()
        try:
            return execute_report(execution, report_execution_actor)
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=1) as pool:
        future = pool.submit(first_attempt)
        assert entered.wait(timeout=10)
        with pytest.raises(ValidationError):
            execute_report(execution, report_execution_actor)
        release.set()
        future.result(timeout=15)

    execution.refresh_from_db()
    assert execution.status == execution.Status.COMPLETED
    assert counts == {'executor': 1, 'store': 1}
    assert (
        ProtectedFile.objects.filter(
            source_model='reports.ReportExecution',
            source_record_id=str(execution.pk),
        ).count()
        == 1
    )


@pytest.mark.parametrize('failure_stage', ('executor', 'renderer', 'storage', 'completion'))
@pytest.mark.django_db
def test_execution_failures_cleanup_artifact_and_persist_sanitized_error(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
    monkeypatch,
    caplog,
    failure_stage,
):
    from files.models import ProtectedFile
    from governance.models import GovernanceAuditLog
    from reports import services

    _configure_report_encryption(settings, tmp_path)
    _registry, executors = isolated_report_registry
    secret = f'segredo-interno-{failure_stage}'

    if failure_stage == 'executor':

        def broken_executor(context):
            raise RuntimeError(secret)

        executors['tests.lifecycle'] = broken_executor
    elif failure_stage == 'renderer':

        def broken_renderer(*args, **kwargs):
            raise RuntimeError(secret)

        monkeypatch.setattr(services, 'render_report', broken_renderer)
    elif failure_stage == 'storage':
        original_store = ProtectedFile.store_encrypted_content

        def broken_store(self, *args, **kwargs):
            original_store(self, *args, **kwargs)
            raise RuntimeError(secret)

        monkeypatch.setattr(ProtectedFile, 'store_encrypted_content', broken_store)
    else:

        def broken_completion(*args, **kwargs):
            raise RuntimeError(secret)

        monkeypatch.setattr(services, '_complete_execution', broken_completion)

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    with pytest.raises(services.ReportExecutionError) as public_error:
        services.execute_report(execution, report_execution_actor)

    assert public_error.value.message_dict == {
        'execution': ['Falha interna ao processar o relatório.'],
    }
    assert secret not in str(public_error.value)
    execution.refresh_from_db()
    assert execution.status == execution.Status.FAILED
    assert execution.completed_at is not None
    assert execution.result_file_id is None
    assert execution.result_reference == ''
    assert execution.content_hash == ''
    assert execution.row_count == 0
    assert execution.error_message == 'Falha interna ao processar o relatório.'
    assert secret not in execution.error_message
    artifacts = ProtectedFile.objects.filter(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    )
    assert all(artifact.status == ProtectedFile.Status.DELETED for artifact in artifacts)
    assert list(tmp_path.rglob('*.enc')) == []
    assert GovernanceAuditLog.objects.filter(
        module='reports',
        target_model='ReportExecution',
        target_record_id=str(execution.pk),
        action='report.execution.failed',
        severity=GovernanceAuditLog.Severity.ERROR,
    ).exists()
    assert secret in caplog.text


@pytest.mark.django_db
def test_validation_failure_is_sanitized_in_pt_br(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
):
    from reports import services

    _registry, executors = isolated_report_registry

    def invalid_executor(context):
        raise ValidationError({'filters': 'detalhe sensível do executor'})

    executors['tests.lifecycle'] = invalid_executor
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    with pytest.raises(services.ReportExecutionError) as public_error:
        services.execute_report(execution, report_execution_actor)

    assert public_error.value.message_dict == {
        'execution': ['Falha de validação ao processar o relatório.'],
    }
    assert 'sensível' not in str(public_error.value)
    execution.refresh_from_db()
    assert execution.status == execution.Status.FAILED
    assert execution.error_message == 'Falha de validação ao processar o relatório.'
    assert 'sensível' not in execution.error_message


@pytest.mark.django_db
def test_cleanup_failure_does_not_mask_original_failure_and_keeps_artifact_hidden(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
    caplog,
):
    from django.core.files.storage import default_storage

    from files.models import ProtectedFile
    from reports import services

    _configure_report_encryption(settings, tmp_path)
    original_store = ProtectedFile.store_encrypted_content

    def broken_store(self, *args, **kwargs):
        original_store(self, *args, **kwargs)
        raise RuntimeError('falha original')

    def broken_delete(name):
        raise OSError('falha cleanup')

    monkeypatch.setattr(ProtectedFile, 'store_encrypted_content', broken_store)
    monkeypatch.setattr(default_storage, 'delete', broken_delete)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    with pytest.raises(services.ReportExecutionError) as public_error:
        services.execute_report(execution, report_execution_actor)

    assert public_error.value.message_dict == {
        'execution': ['Falha interna ao processar o relatório.'],
    }
    assert 'falha original' not in str(public_error.value)
    execution.refresh_from_db()
    artifact = ProtectedFile.objects.get(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    )
    assert execution.status == execution.Status.FAILED
    assert execution.result_file_id == artifact.pk
    assert artifact.status == artifact.Status.ACTIVE
    with pytest.raises(ValidationError):
        artifact.read_encrypted_content(report_execution_actor)
    assert 'falha cleanup' in caplog.text


@pytest.mark.django_db
def test_lost_execution_ownership_cleans_artifact_without_overwriting_new_state(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
):
    from files.models import ProtectedFile
    from reports.contracts import ReportColumn, ReportDataset
    from reports.models import ReportExecution
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    _registry, executors = isolated_report_registry
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    def ownership_losing_executor(context):
        ReportExecution.objects.filter(pk=execution.pk).update(
            status=ReportExecution.Status.CANCELLED,
            completed_at=timezone.now(),
        )
        return ReportDataset(
            title='Posse perdida',
            columns=(ReportColumn('value', 'Valor'),),
            rows=({'value': 'linha'},),
        )

    executors['tests.lifecycle'] = ownership_losing_executor

    with pytest.raises(ValidationError) as invalid:
        execute_report(execution, report_execution_actor)

    assert invalid.value.message_dict == {
        'execution': ['Falha de validação ao processar o relatório.'],
    }
    execution.refresh_from_db()
    assert execution.status == execution.Status.CANCELLED
    assert execution.result_file_id is None
    assert not ProtectedFile.objects.filter(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    ).exists()
    assert list(tmp_path.rglob('*.enc')) == []


@pytest.mark.django_db
def test_notification_failure_keeps_durable_completion_and_logs_warning(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from governance.models import GovernanceAuditLog
    from reports.models import ReportExecution
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)

    def broken_notification(self):
        raise RuntimeError('notificação indisponível')

    monkeypatch.setattr(ReportExecution, 'notify_completion', broken_notification)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    completed = execute_report(execution, report_execution_actor)

    completed.refresh_from_db()
    assert completed.status == completed.Status.COMPLETED
    assert completed.result_file_id is not None
    assert GovernanceAuditLog.objects.filter(
        module='reports',
        target_model='ReportExecution',
        target_record_id=str(completed.pk),
        action='report.execution.notification_failed',
        severity=GovernanceAuditLog.Severity.WARNING,
    ).exists()


@pytest.mark.django_db
def test_result_file_protects_completed_execution_evidence(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from django.db.models.deletion import ProtectedError

    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    completed = execute_report(execution, report_execution_actor)

    with pytest.raises(ProtectedError):
        completed.result_file.delete()


@pytest.mark.django_db
def test_report_execution_run_and_celery_task_delegate_with_persisted_requester(
    system_report_definition,
    report_execution_actor,
    monkeypatch,
):
    from reports import services
    from reports.models import ReportExecution
    from reports.tasks import generate_report_execution

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    calls = []

    def fake_execute(target, actor):
        calls.append(('service', target.pk, actor.pk))
        return target

    monkeypatch.setattr(services, 'execute_report', fake_execute)
    assert execution.run() == execution
    assert calls == [('service', execution.pk, report_execution_actor.pk)]

    calls.clear()

    def fake_run(self, user=None):
        calls.append(('task', self.pk, user))
        return self

    monkeypatch.setattr(ReportExecution, 'run', fake_run)
    assert generate_report_execution(execution.pk) == {
        'execution_id': execution.pk,
        'status': execution.status,
    }
    assert calls == [('task', execution.pk, None)]


@pytest.mark.parametrize(
    ('field_name', 'replacement'),
    (
        ('filters', {'status': 'tampered'}),
        ('export_format', 'pdf'),
        ('execution_number', 'REP-TAMPERED'),
        ('requested_at', datetime(2026, 7, 29, tzinfo=timezone.get_current_timezone())),
    ),
)
@pytest.mark.django_db
def test_report_execution_inputs_are_immutable_via_model_clean_and_save(
    system_report_definition,
    report_execution_actor,
    field_name,
    replacement,
):
    execution = system_report_definition.create_execution(
        filters={'status': 'original'},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    setattr(execution, field_name, replacement)

    with pytest.raises(ValidationError) as invalid_clean:
        execution.full_clean()
    assert field_name in invalid_clean.value.message_dict

    with pytest.raises(ValidationError) as invalid_save:
        execution.save()
    assert field_name in invalid_save.value.message_dict


@pytest.mark.django_db
def test_report_execution_relationship_inputs_are_immutable_after_creation(
    system_report_definition,
    report_execution_actor,
    django_user_model,
):
    from reports.models import ReportDefinition, ReportExecution, ReportSchedule

    other_user = django_user_model.objects.create_user(
        username='immutable.other@example.com',
        email='immutable.other@example.com',
    )
    other_definition = ReportDefinition.objects.create(
        code='USR-IMMUTABLE-OTHER',
        title='Outra definição',
        module=ReportDefinition.Module.QUALITY,
        category=ReportDefinition.Category.OPERATIONAL,
        allowed_export_formats=[ReportExecution.ExportFormat.CSV],
    )
    schedule = ReportSchedule.objects.create(
        definition=system_report_definition,
        name='Agenda imutável',
        frequency=ReportSchedule.Frequency.DAILY,
        filters={},
        export_format=ReportExecution.ExportFormat.CSV,
        next_run_at=timezone.now() + timedelta(days=1),
        owner=report_execution_actor,
    )

    replacements = {
        'definition': other_definition,
        'requested_by': other_user,
        'schedule': schedule,
    }
    for field_name, replacement in replacements.items():
        execution = system_report_definition.create_execution(
            filters={},
            export_format='csv',
            requested_by=report_execution_actor,
        )
        setattr(execution, field_name, replacement)
        with pytest.raises(ValidationError) as invalid:
            execution.save()
        assert field_name in invalid.value.message_dict


@pytest.mark.parametrize(
    'field_name',
    (
        'definition',
        'filters',
        'export_format',
        'requested_by',
        'schedule',
        'execution_number',
        'requested_at',
    ),
)
@pytest.mark.django_db
def test_report_execution_serializer_explicitly_rejects_persisted_input_updates(
    system_report_definition,
    report_execution_actor,
    field_name,
):
    from reports.serializers import ReportExecutionSerializer

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    serializer = ReportExecutionSerializer(
        execution,
        data={field_name: 'attempted-change'},
        partial=True,
    )

    assert serializer.is_valid() is False
    assert str(serializer.errors[field_name]) == (
        'Este dado de entrada não pode ser alterado após criar a execução.'
    )


@pytest.mark.parametrize('status', ('pending', 'running', 'completed'))
@pytest.mark.django_db
def test_report_execution_api_rejects_input_patch_in_every_lifecycle_state(
    system_report_definition,
    report_execution_actor,
    status,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution

    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='change_reportexecution',
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    ReportExecution.objects.filter(pk=execution.pk).update(status=status)
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.patch(
        f'/api/reports/executions/{execution.pk}/',
        {'filters': {'status': 'tampered'}},
        format='json',
    )

    assert response.status_code == 400
    assert response.json()['filters'] == (
        'Este dado de entrada não pode ser alterado após criar a execução.'
    )
    execution.refresh_from_db()
    assert execution.filters == {}


@pytest.mark.django_db
def test_report_execution_rejects_hostile_filter_object_without_invoking_dunders(
    system_report_definition,
    report_execution_actor,
):
    class HostileDict(dict):
        def __bool__(self):
            raise AssertionError('dunder de filtros não pode ser chamado')

        def __iter__(self):
            raise AssertionError('dunder de filtros não pode ser chamado')

        def items(self):
            raise AssertionError('dunder de filtros não pode ser chamado')

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    execution.filters = HostileDict()

    with pytest.raises(ValidationError) as invalid:
        execution.full_clean()

    assert invalid.value.message_dict == {
        'filters': ['Filtros da execução devem ser um objeto JSON seguro.'],
    }


@pytest.mark.django_db
def test_query_update_tamper_during_render_prevents_completion_and_cleans_artifact(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
):
    from files.models import ProtectedFile
    from reports.contracts import ReportColumn, ReportDataset
    from reports.models import ReportExecution
    from reports.services import ReportExecutionError, execute_report

    _configure_report_encryption(settings, tmp_path)
    _registry, executors = isolated_report_registry
    execution = system_report_definition.create_execution(
        filters={'status': 'original'},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    observed_filters = []

    def tampering_executor(context):
        observed_filters.append(dict(context.filters))
        ReportExecution.objects.filter(pk=execution.pk).update(filters={'status': 'tampered'})
        return ReportDataset(
            title='Snapshot',
            columns=(ReportColumn('value', 'Valor'),),
            rows=({'value': 'linha'},),
        )

    executors['tests.lifecycle'] = tampering_executor

    with pytest.raises(ReportExecutionError) as public_error:
        execute_report(execution, report_execution_actor)

    assert public_error.value.message_dict == {
        'execution': ['Falha de validação ao processar o relatório.'],
    }
    assert observed_filters == [{'status': 'original'}]
    execution.refresh_from_db()
    assert execution.status == execution.Status.FAILED
    assert execution.filters == {'status': 'tampered'}
    assert execution.result_file_id is None
    assert not ProtectedFile.objects.filter(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    ).exists()
    assert list(tmp_path.rglob('*.enc')) == []


@pytest.mark.django_db
def test_execution_run_api_never_exposes_executor_details(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    caplog,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    _registry, executors = isolated_report_registry
    secret = 'segredo-do-conector-externo'

    def broken_executor(context):
        raise RuntimeError(secret)

    executors['tests.lifecycle'] = broken_executor
    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='change_reportexecution',
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.post(f'/api/reports/executions/{execution.pk}/run/')

    assert response.status_code == 400
    assert response.json() == {
        'execution': ['Falha interna ao processar o relatório.'],
    }
    assert secret not in response.content.decode()
    execution.refresh_from_db()
    assert execution.error_message == 'Falha interna ao processar o relatório.'
    assert secret not in execution.error_message
    assert secret in caplog.text


@pytest.mark.parametrize(
    'unsafe_reference',
    (
        '/protected/execution/file.enc',
        'protected/../file.enc',
        'protected/execution/../../file.enc',
        'protected\\execution\\file.enc',
        'protected/execution/file.enc\x00suffix',
        'protected//file.enc',
        'protected/./file.enc',
    ),
)
@pytest.mark.django_db
def test_cleanup_never_deletes_noncanonical_storage_references(
    report_execution_actor,
    monkeypatch,
    unsafe_reference,
):
    from django.core.files.storage import default_storage

    from files.models import ProtectedFile
    from reports.services import _cleanup_artifact

    delete_attempts = []
    monkeypatch.setattr(
        default_storage,
        'delete',
        lambda reference: delete_attempts.append(reference),
    )
    artifact = ProtectedFile(
        pk=999_999,
        file_reference=unsafe_reference,
    )

    _cleanup_artifact(artifact, report_execution_actor)

    assert delete_attempts == []


@pytest.mark.django_db
def test_completed_execution_is_idempotent_without_second_render_or_artifact(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
):
    from files.models import ProtectedFile
    from reports.contracts import ReportColumn, ReportDataset
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    _registry, executors = isolated_report_registry
    calls = []

    def counted_executor(context):
        calls.append(dict(context.filters))
        return ReportDataset(
            title='Idempotência',
            columns=(ReportColumn('value', 'Valor'),),
            rows=({'value': 'única'},),
        )

    executors['tests.lifecycle'] = counted_executor
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    first = execute_report(execution, report_execution_actor)
    second = execute_report(execution, report_execution_actor)

    assert second.pk == first.pk
    assert second.result_file_id == first.result_file_id
    assert calls == [{}]
    assert (
        ProtectedFile.objects.filter(
            source_model='reports.ReportExecution',
            source_record_id=str(execution.pk),
            status=ProtectedFile.Status.ACTIVE,
        ).count()
        == 1
    )


@pytest.mark.django_db
def test_fresh_running_execution_has_typed_lease_conflict_and_api_409(
    system_report_definition,
    report_execution_actor,
    settings,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution
    from reports.services import ReportExecutionInProgress, execute_report

    settings.REPORT_EXECUTION_LEASE_SECONDS = 300
    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='change_reportexecution',
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    ReportExecution.objects.filter(pk=execution.pk).update(
        status=ReportExecution.Status.RUNNING,
        started_at=timezone.now(),
    )

    with pytest.raises(ReportExecutionInProgress) as conflict:
        execute_report(execution, report_execution_actor)
    assert 1 <= conflict.value.retry_after <= 300

    client = APIClient()
    client.force_authenticate(report_execution_actor)
    response = client.post(f'/api/reports/executions/{execution.pk}/run/')
    assert response.status_code == 409
    assert response.json() == {
        'execution': ['A execução do relatório já está em andamento.'],
    }


@pytest.mark.django_db
def test_stale_running_execution_is_reclaimed_and_completed(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from reports.models import ReportExecution
    from reports.services import execute_report

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 60
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    stale_started_at = timezone.now() - timedelta(seconds=61)
    ReportExecution.objects.filter(pk=execution.pk).update(
        status=ReportExecution.Status.RUNNING,
        started_at=stale_started_at,
    )

    completed = execute_report(execution, report_execution_actor)

    assert completed.status == completed.Status.COMPLETED
    assert completed.started_at > stale_started_at
    assert completed.result_file_id is not None


@pytest.mark.django_db
def test_report_task_is_worker_loss_safe_and_second_delivery_is_idempotent(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from reports.tasks import generate_report_execution

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    first = generate_report_execution.run(execution.pk)
    second = generate_report_execution.run(execution.pk)

    assert generate_report_execution.acks_late is True
    assert generate_report_execution.reject_on_worker_lost is True
    assert generate_report_execution.max_retries == 5
    assert first == {'execution_id': execution.pk, 'status': 'completed'}
    assert second == first


def test_report_task_public_exceptions_are_serializable_for_worker_transport():
    import pickle

    from reports.services import (
        ReportExecutionError,
        ReportExecutionInProgress,
        ReportExecutionRetryableError,
    )

    exceptions = (
        ReportExecutionError(),
        ReportExecutionInProgress(42),
        ReportExecutionRetryableError(),
    )
    restored = [pickle.loads(pickle.dumps(error)) for error in exceptions]

    assert [type(error) for error in restored] == [type(error) for error in exceptions]
    assert restored[1].retry_after == 42
    assert all('segredo' not in str(error) for error in restored)


@pytest.mark.django_db
def test_report_task_retries_fresh_lease_with_countdown(
    system_report_definition,
    report_execution_actor,
    monkeypatch,
):
    from reports.models import ReportExecution
    from reports.tasks import generate_report_execution

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    ReportExecution.objects.filter(pk=execution.pk).update(
        status=ReportExecution.Status.RUNNING,
        started_at=timezone.now(),
    )
    retry_call = {}

    class RetryRequested(Exception):
        pass

    def fake_retry(**kwargs):
        retry_call.update(kwargs)
        raise RetryRequested

    monkeypatch.setattr(generate_report_execution, 'retry', fake_retry)

    with pytest.raises(RetryRequested):
        generate_report_execution.run(execution.pk)

    assert retry_call['countdown'] > 0
    assert 'em andamento' in str(retry_call['exc'])


@pytest.mark.parametrize(
    ('raised_error', 'expected_status', 'expects_retry'),
    (
        (OSError('storage indisponível'), 'pending', True),
        (ValidationError({'filters': 'regra inválida'}), 'failed', False),
    ),
)
@pytest.mark.django_db
def test_report_task_retries_only_transient_execution_failures(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    monkeypatch,
    raised_error,
    expected_status,
    expects_retry,
):
    from reports.services import ReportExecutionError
    from reports.tasks import generate_report_execution

    _registry, executors = isolated_report_registry

    def broken_executor(context):
        raise raised_error

    executors['tests.lifecycle'] = broken_executor
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    class RetryRequested(Exception):
        pass

    retry_call = {}

    def fake_retry(**kwargs):
        retry_call.update(kwargs)
        raise RetryRequested

    monkeypatch.setattr(generate_report_execution, 'retry', fake_retry)
    expected_exception = RetryRequested if expects_retry else ReportExecutionError
    with pytest.raises(expected_exception):
        generate_report_execution.run(execution.pk)

    execution.refresh_from_db()
    assert execution.status == expected_status
    assert bool(retry_call) is expects_retry


@pytest.mark.django_db(transaction=True)
def test_reclaimed_worker_owns_only_final_active_artifact(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
):
    from threading import Event, Lock

    from django.db import close_old_connections

    from files.models import ProtectedFile
    from reports.contracts import ReportColumn, ReportDataset
    from reports.models import ReportExecution
    from reports.services import ReportExecutionError, execute_report

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 30
    _registry, executors = isolated_report_registry
    first_entered = Event()
    release_first = Event()
    lock = Lock()
    invocation = 0

    def racing_executor(context):
        nonlocal invocation
        with lock:
            invocation += 1
            current = invocation
        if current == 1:
            first_entered.set()
            assert release_first.wait(timeout=10)
        return ReportDataset(
            title='Lease',
            columns=(ReportColumn('worker', 'Worker'),),
            rows=({'worker': current},),
        )

    executors['tests.lifecycle'] = racing_executor
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )

    def old_worker():
        close_old_connections()
        try:
            return execute_report(execution, report_execution_actor)
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=1) as pool:
        old_result = pool.submit(old_worker)
        assert first_entered.wait(timeout=10)
        ReportExecution.objects.filter(pk=execution.pk).update(
            started_at=timezone.now() - timedelta(seconds=31),
        )
        new_result = execute_report(execution, report_execution_actor)
        release_first.set()
        with pytest.raises(ReportExecutionError):
            old_result.result(timeout=15)

    execution.refresh_from_db()
    artifacts = ProtectedFile.objects.filter(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    )
    assert new_result.pk == execution.pk
    assert execution.status == execution.Status.COMPLETED
    assert artifacts.filter(status=ProtectedFile.Status.ACTIVE).count() == 1
    assert artifacts.filter(status=ProtectedFile.Status.DELETED).count() == 0
    assert execution.result_file.status == ProtectedFile.Status.ACTIVE


@pytest.mark.django_db
def test_worker_death_after_reservation_is_reconciled_before_stale_reclaim(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from files.models import ProtectedFile
    from reports import services
    from reports.models import ReportExecution

    class SimulatedWorkerDeath(BaseException):
        pass

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 30
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    original_store = services._store_protected_artifact

    def die_before_store(*args, **kwargs):
        raise SimulatedWorkerDeath

    monkeypatch.setattr(services, '_store_protected_artifact', die_before_store)
    with pytest.raises(SimulatedWorkerDeath):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    abandoned_id = execution.result_file_id
    assert execution.status == ReportExecution.Status.RUNNING
    assert abandoned_id is not None
    abandoned = ProtectedFile.objects.get(pk=abandoned_id)
    assert abandoned.file_reference.startswith(f'protected/{abandoned.file_number}/')
    assert abandoned.file_reference.endswith('.enc')
    assert list(tmp_path.rglob('*.enc')) == []

    ReportExecution.objects.filter(pk=execution.pk).update(
        started_at=timezone.now() - timedelta(seconds=31),
    )
    monkeypatch.setattr(services, '_store_protected_artifact', original_store)
    completed = services.execute_report(execution, report_execution_actor)

    abandoned.refresh_from_db()
    assert completed.status == completed.Status.COMPLETED
    assert completed.result_file_id != abandoned_id
    assert abandoned.status == ProtectedFile.Status.DELETED
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / completed.result_file.file_reference]


@pytest.mark.django_db
def test_worker_death_after_real_store_is_reconciled_without_orphan_blob(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from files.models import ProtectedFile
    from reports import services
    from reports.models import ReportExecution

    class SimulatedWorkerDeath(BaseException):
        pass

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 30
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    original_complete = services._complete_execution

    def die_before_completion(*args, **kwargs):
        raise SimulatedWorkerDeath

    monkeypatch.setattr(services, '_complete_execution', die_before_completion)
    with pytest.raises(SimulatedWorkerDeath):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    abandoned_id = execution.result_file_id
    abandoned = ProtectedFile.objects.get(pk=abandoned_id)
    abandoned_blob = tmp_path / abandoned.file_reference
    assert execution.status == ReportExecution.Status.RUNNING
    assert abandoned_blob.exists()

    ReportExecution.objects.filter(pk=execution.pk).update(
        started_at=timezone.now() - timedelta(seconds=31),
    )
    monkeypatch.setattr(services, '_complete_execution', original_complete)
    completed = services.execute_report(execution, report_execution_actor)

    abandoned.refresh_from_db()
    assert completed.result_file_id != abandoned_id
    assert abandoned.status == ProtectedFile.Status.DELETED
    assert abandoned_blob.exists() is False
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / completed.result_file.file_reference]


@pytest.mark.django_db(transaction=True)
def test_old_reserved_worker_cannot_clear_or_delete_new_worker_artifact(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from threading import Event

    from django.db import close_old_connections

    from files.models import ProtectedFile
    from reports import services
    from reports.models import ReportExecution

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 30
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    old_reserved = Event()
    release_old = Event()
    original_store = services._store_protected_artifact
    first_call = True

    def blocking_store(*args, **kwargs):
        nonlocal first_call
        if first_call:
            first_call = False
            old_reserved.set()
            assert release_old.wait(timeout=10)
        return original_store(*args, **kwargs)

    monkeypatch.setattr(services, '_store_protected_artifact', blocking_store)

    def old_worker():
        close_old_connections()
        try:
            return services.execute_report(execution, report_execution_actor)
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=1) as pool:
        old_result = pool.submit(old_worker)
        assert old_reserved.wait(timeout=10)
        execution.refresh_from_db()
        old_artifact_id = execution.result_file_id
        ReportExecution.objects.filter(pk=execution.pk).update(
            started_at=timezone.now() - timedelta(seconds=31),
        )
        new_result = services.execute_report(execution, report_execution_actor)
        release_old.set()
        with pytest.raises(services.ReportExecutionError):
            old_result.result(timeout=15)

    execution.refresh_from_db()
    old_artifact = ProtectedFile.objects.get(pk=old_artifact_id)
    assert execution.status == execution.Status.COMPLETED
    assert execution.result_file_id == new_result.result_file_id
    assert execution.result_file_id != old_artifact_id
    assert execution.result_file.status == ProtectedFile.Status.ACTIVE
    assert old_artifact.status == ProtectedFile.Status.DELETED
    assert (
        ProtectedFile.objects.filter(
            source_model='reports.ReportExecution',
            source_record_id=str(execution.pk),
            status=ProtectedFile.Status.ACTIVE,
        ).count()
        == 1
    )
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / execution.result_file.file_reference]


@pytest.mark.django_db(transaction=True)
def test_upload_holds_reservation_locks_and_cannot_leave_post_reclaim_blob(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from threading import Event, Lock

    from django.db import close_old_connections

    from files.models import ProtectedFile
    from reports import services
    from reports.models import ReportExecution

    class SimulatedWorkerDeath(BaseException):
        pass

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 30
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    upload_entered = Event()
    release_upload = Event()
    stale_update_completed = Event()
    original_store = ProtectedFile.store_encrypted_content
    store_lock = Lock()
    first_store = True

    def paused_store(self, *args, **kwargs):
        nonlocal first_store
        with store_lock:
            is_old_worker = first_store
            first_store = False
        if not is_old_worker:
            return original_store(self, *args, **kwargs)
        upload_entered.set()
        assert release_upload.wait(timeout=10)
        original_store(self, *args, **kwargs)
        raise SimulatedWorkerDeath

    monkeypatch.setattr(ProtectedFile, 'store_encrypted_content', paused_store)

    def old_worker():
        close_old_connections()
        try:
            return services.execute_report(execution, report_execution_actor)
        finally:
            close_old_connections()

    def reclaim_worker():
        close_old_connections()
        try:
            ReportExecution.objects.filter(pk=execution.pk).update(
                started_at=timezone.now() - timedelta(seconds=31),
            )
            stale_update_completed.set()
            return services.execute_report(execution, report_execution_actor)
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=2) as pool:
        old_result = pool.submit(old_worker)
        assert upload_entered.wait(timeout=10)
        reclaim_result = pool.submit(reclaim_worker)
        try:
            assert stale_update_completed.wait(timeout=2) is False
        finally:
            release_upload.set()
        with pytest.raises(SimulatedWorkerDeath):
            old_result.result(timeout=15)
        completed = reclaim_result.result(timeout=15)

    execution.refresh_from_db()
    artifacts = ProtectedFile.objects.filter(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    )
    assert execution.status == execution.Status.COMPLETED
    assert execution.result_file_id == completed.result_file_id
    assert artifacts.filter(status=ProtectedFile.Status.ACTIVE).count() == 1
    assert artifacts.filter(status=ProtectedFile.Status.DELETED).count() == 1
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / completed.result_file.file_reference]


@pytest.mark.django_db
def test_reserved_report_artifact_is_hidden_until_execution_completes(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from files.models import ProtectedFile
    from reports import services

    class SimulatedWorkerDeath(BaseException):
        pass

    _configure_report_encryption(settings, tmp_path)
    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='view_reportexecution',
        ),
        Permission.objects.get(
            content_type__app_label='files',
            codename='view_protectedfile',
        ),
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    monkeypatch.setattr(
        services,
        '_store_protected_artifact',
        lambda *args, **kwargs: (_ for _ in ()).throw(SimulatedWorkerDeath),
    )
    with pytest.raises(SimulatedWorkerDeath):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    reserved = ProtectedFile.objects.get(pk=execution.result_file_id)
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    execution_response = client.get(f'/api/reports/executions/{execution.pk}/')
    file_response = client.get(f'/api/files/protected-files/{reserved.pk}/')

    assert execution_response.status_code == 200
    assert execution_response.json()['result_file'] is None
    assert 'result_reference' not in execution_response.json()
    assert execution_response.json()['content_hash'] == ''
    assert file_response.status_code == 404
    with pytest.raises(ValidationError) as unavailable:
        reserved.generate_secure_link(report_execution_actor)
    assert unavailable.value.message_dict == {
        'status': ['Arquivo de relatório ainda não foi concluído.'],
    }


@pytest.mark.django_db
def test_stale_reclaim_cleanup_failure_returns_pending_before_new_render(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    settings,
    tmp_path,
    monkeypatch,
):
    from django.core.files.storage import default_storage

    from files.models import ProtectedFile
    from governance.models import GovernanceAuditLog
    from reports import services
    from reports.contracts import ReportColumn, ReportDataset
    from reports.models import ReportExecution

    class SimulatedWorkerDeath(BaseException):
        pass

    _configure_report_encryption(settings, tmp_path)
    settings.REPORT_EXECUTION_LEASE_SECONDS = 30
    _registry, executors = isolated_report_registry
    renders = []

    def counted_executor(context):
        renders.append(dict(context.filters))
        return ReportDataset(
            title='Reconciliação',
            columns=(ReportColumn('value', 'Valor'),),
            rows=({'value': len(renders)},),
        )

    executors['tests.lifecycle'] = counted_executor
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    original_complete = services._complete_execution
    monkeypatch.setattr(
        services,
        '_complete_execution',
        lambda *args, **kwargs: (_ for _ in ()).throw(SimulatedWorkerDeath),
    )
    with pytest.raises(SimulatedWorkerDeath):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    abandoned = ProtectedFile.objects.get(pk=execution.result_file_id)
    abandoned_blob = tmp_path / abandoned.file_reference
    ReportExecution.objects.filter(pk=execution.pk).update(
        started_at=timezone.now() - timedelta(seconds=31),
    )
    original_delete = default_storage.delete

    def unavailable_delete(reference):
        raise OSError('storage indisponível')

    monkeypatch.setattr(default_storage, 'delete', unavailable_delete)
    monkeypatch.setattr(services, '_complete_execution', original_complete)
    with pytest.raises(services.ReportExecutionRetryableError):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    abandoned.refresh_from_db()
    assert execution.status == execution.Status.PENDING
    assert execution.result_file_id == abandoned.pk
    assert abandoned.status == ProtectedFile.Status.ACTIVE
    assert abandoned_blob.exists()
    assert renders == [{}]

    monkeypatch.setattr(default_storage, 'delete', original_delete)
    completed = services.execute_report(execution, report_execution_actor)

    abandoned.refresh_from_db()
    assert completed.status == completed.Status.COMPLETED
    assert abandoned.status == ProtectedFile.Status.DELETED
    assert abandoned_blob.exists() is False
    assert renders == [{}, {}]
    actions = GovernanceAuditLog.objects.filter(
        module='reports',
        target_model='ReportExecution',
        target_record_id=str(execution.pk),
    ).values_list('action', flat=True)
    assert 'report.execution.retry_scheduled' in actions
    assert 'report.execution.reclaimed' in actions
    assert 'report.execution.completed' in actions


@pytest.mark.django_db
def test_transient_cleanup_failure_keeps_artifact_recoverable_for_next_retry(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from django.contrib.auth.models import Permission
    from django.core.files.storage import default_storage
    from rest_framework.test import APIClient

    from files.models import ProtectedFile
    from reports import services

    _configure_report_encryption(settings, tmp_path)
    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='view_reportexecution',
        ),
        Permission.objects.get(
            content_type__app_label='files',
            codename='view_protectedfile',
        ),
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    original_store = ProtectedFile.store_encrypted_content
    original_delete = default_storage.delete

    def store_then_fail(self, *args, **kwargs):
        original_store(self, *args, **kwargs)
        raise OSError('conexão caiu após o upload')

    def cleanup_unavailable(reference):
        raise OSError('storage indisponível para cleanup')

    monkeypatch.setattr(ProtectedFile, 'store_encrypted_content', store_then_fail)
    monkeypatch.setattr(default_storage, 'delete', cleanup_unavailable)
    with pytest.raises(services.ReportExecutionRetryableError):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    abandoned = ProtectedFile.objects.get(
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
    )
    abandoned_blob = tmp_path / abandoned.file_reference
    assert execution.status == execution.Status.PENDING
    assert execution.result_file_id == abandoned.pk
    assert abandoned.status == abandoned.Status.ACTIVE
    assert abandoned_blob.exists()
    client = APIClient()
    client.force_authenticate(report_execution_actor)
    assert client.get(f'/api/files/protected-files/{abandoned.pk}/').status_code == 404
    with pytest.raises(ValidationError) as unavailable:
        abandoned.read_encrypted_content(report_execution_actor)
    assert unavailable.value.message_dict == {
        'status': ['Arquivo de relatório ainda não foi concluído.'],
    }

    monkeypatch.setattr(ProtectedFile, 'store_encrypted_content', original_store)
    monkeypatch.setattr(default_storage, 'delete', original_delete)
    completed = services.execute_report(execution, report_execution_actor)

    abandoned.refresh_from_db()
    assert completed.status == completed.Status.COMPLETED
    assert abandoned.status == abandoned.Status.DELETED
    assert abandoned_blob.exists() is False
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / completed.result_file.file_reference]


@pytest.mark.django_db
def test_storage_rename_cleanup_failure_is_reconciled_from_reserved_directory(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from django.core.files.storage import default_storage

    from files.models import ProtectedFile
    from reports import services

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    original_save = default_storage.save
    original_delete = default_storage.delete
    alternate_cleanup_attempts = 0

    def renamed_save(name, content, max_length=None):
        alternate = name.removesuffix('.enc') + '_renamed.enc'
        return original_save(alternate, content, max_length=max_length)

    def unavailable_alternate_cleanup(name):
        nonlocal alternate_cleanup_attempts
        if name.endswith('_renamed.enc'):
            alternate_cleanup_attempts += 1
        if name.endswith('_renamed.enc') and alternate_cleanup_attempts <= 2:
            raise OSError('cleanup alternativo indisponível')
        return original_delete(name)

    monkeypatch.setattr(default_storage, 'save', renamed_save)
    monkeypatch.setattr(default_storage, 'delete', unavailable_alternate_cleanup)

    with pytest.raises(services.ReportExecutionRetryableError):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    abandoned = ProtectedFile.objects.get(pk=execution.result_file_id)
    alternate_blobs = list(tmp_path.rglob('*_renamed.enc'))
    assert execution.status == execution.Status.PENDING
    assert abandoned.status == ProtectedFile.Status.ACTIVE
    assert len(alternate_blobs) == 1
    assert alternate_blobs[0].parent == tmp_path / f'protected/{abandoned.file_number}'

    monkeypatch.setattr(default_storage, 'save', original_save)
    monkeypatch.setattr(default_storage, 'delete', original_delete)
    completed = services.execute_report(execution, report_execution_actor)

    abandoned.refresh_from_db()
    assert completed.status == completed.Status.COMPLETED
    assert completed.result_file_id != abandoned.pk
    assert abandoned.status == ProtectedFile.Status.DELETED
    assert alternate_blobs[0].exists() is False
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / completed.result_file.file_reference]


@pytest.mark.django_db
def test_storage_rename_with_immediate_cleanup_remains_retryable(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from django.core.files.storage import default_storage

    from reports import services

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    original_save = default_storage.save

    def renamed_save(name, content, max_length=None):
        alternate = name.removesuffix('.enc') + '_renamed.enc'
        return original_save(alternate, content, max_length=max_length)

    monkeypatch.setattr(default_storage, 'save', renamed_save)
    with pytest.raises(services.ReportExecutionRetryableError):
        services.execute_report(execution, report_execution_actor)

    execution.refresh_from_db()
    assert execution.status == execution.Status.PENDING
    assert execution.result_file_id is None
    assert list(tmp_path.rglob('*.enc')) == []

    monkeypatch.setattr(default_storage, 'save', original_save)
    completed = services.execute_report(execution, report_execution_actor)

    assert completed.status == completed.Status.COMPLETED
    assert list(tmp_path.rglob('*.enc')) == [tmp_path / completed.result_file.file_reference]


@pytest.mark.django_db
def test_report_task_marks_pending_execution_failed_when_retry_budget_is_exhausted(
    system_report_definition,
    report_execution_actor,
    isolated_report_registry,
    monkeypatch,
):
    from governance.models import GovernanceAuditLog
    from reports.services import ReportExecutionRetryableError
    from reports.tasks import generate_report_execution

    _registry, executors = isolated_report_registry
    executor_calls = []

    def unavailable_executor(context):
        executor_calls.append(dict(context.filters))
        raise OSError('segredo do storage externo')

    executors['tests.lifecycle'] = unavailable_executor
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    retry_calls = []

    class RetryRequested(Exception):
        pass

    def fake_retry(**kwargs):
        retry_calls.append(kwargs)
        raise RetryRequested

    monkeypatch.setattr(generate_report_execution, 'retry', fake_retry)
    for retry_number in range(generate_report_execution.max_retries + 1):
        generate_report_execution.push_request(retries=retry_number)
        try:
            expected = (
                RetryRequested
                if retry_number < generate_report_execution.max_retries
                else ReportExecutionRetryableError
            )
            with pytest.raises(expected):
                generate_report_execution.run(execution.pk)
        finally:
            generate_report_execution.pop_request()

    execution.refresh_from_db()
    assert len(executor_calls) == generate_report_execution.max_retries + 1
    assert len(retry_calls) == generate_report_execution.max_retries
    assert execution.status == execution.Status.FAILED
    assert execution.completed_at is not None
    assert execution.result_file_id is None
    assert execution.error_message == ('Falha temporária após esgotar as tentativas do relatório.')
    assert 'segredo' not in execution.error_message
    assert (
        GovernanceAuditLog.objects.filter(
            module='reports',
            target_model='ReportExecution',
            target_record_id=str(execution.pk),
            action='report.execution.retry_exhausted',
        ).count()
        == 1
    )


@pytest.mark.django_db
def test_retry_exhaustion_keeps_uncleanable_reservation_linked_and_hidden(
    system_report_definition,
    report_execution_actor,
):
    from files.models import ProtectedFile
    from reports.services import mark_retry_exhausted

    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    reserved = ProtectedFile.objects.create(
        source_module=ProtectedFile.SourceModule.OPERATIONAL,
        source_model='reports.ReportExecution',
        source_record_id=str(execution.pk),
        file_type=ProtectedFile.FileType.REPORT,
        origin=ProtectedFile.Origin.SYSTEM,
        criticality=ProtectedFile.Criticality.MEDIUM,
        confidentiality=ProtectedFile.Confidentiality.INTERNAL,
        title='Reserva com cleanup indisponível',
        file_name='reserva.csv',
        file_reference='protected/reserved/retry.enc',
        mime_type='text/csv',
        file_size=0,
        content_hash='sha256:pending',
        responsible=report_execution_actor,
        uploaded_by=report_execution_actor,
    )
    execution.result_file = reserved
    execution.save(update_fields=['result_file', 'updated_at'])

    assert mark_retry_exhausted(execution.pk) is True

    execution.refresh_from_db()
    reserved.refresh_from_db()
    assert execution.status == execution.Status.FAILED
    assert execution.result_file_id == reserved.pk
    assert reserved.status == ProtectedFile.Status.ACTIVE
    with pytest.raises(ValidationError):
        reserved.read_encrypted_content(report_execution_actor)


@pytest.mark.django_db
def test_retry_exhaustion_never_overwrites_completed_or_active_running_execution(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    monkeypatch,
):
    from reports.models import ReportExecution
    from reports.tasks import generate_report_execution

    _configure_report_encryption(settings, tmp_path)
    completed_execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    completed = generate_report_execution.run(completed_execution.pk)

    def retry_must_not_run(**kwargs):
        raise AssertionError('entrega idempotente não deve agendar retry')

    monkeypatch.setattr(generate_report_execution, 'retry', retry_must_not_run)
    generate_report_execution.push_request(retries=generate_report_execution.max_retries)
    try:
        assert generate_report_execution.run(completed_execution.pk) == completed
    finally:
        generate_report_execution.pop_request()

    running_execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    running_started_at = timezone.now()
    ReportExecution.objects.filter(pk=running_execution.pk).update(
        status=ReportExecution.Status.RUNNING,
        started_at=running_started_at,
    )
    retry_calls = []

    class RetryRequested(Exception):
        pass

    def capture_retry(**kwargs):
        retry_calls.append(kwargs)
        raise RetryRequested

    monkeypatch.setattr(generate_report_execution, 'retry', capture_retry)
    generate_report_execution.push_request(retries=generate_report_execution.max_retries)
    try:
        with pytest.raises(RetryRequested):
            generate_report_execution.run(running_execution.pk)
    finally:
        generate_report_execution.pop_request()

    running_execution.refresh_from_db()
    assert retry_calls
    assert running_execution.status == running_execution.Status.RUNNING
    assert running_execution.started_at == running_started_at


CURATED_REPORT_EXPECTATIONS = {
    'REL-FIN-001': (
        'finance.receivables_open_overdue',
        'Contas a receber em aberto e vencidas',
        'finance',
        'operational',
        (),
    ),
    'REL-FIN-002': (
        'finance.payables_open_overdue',
        'Contas a pagar em aberto e vencidas',
        'finance',
        'operational',
        (),
    ),
    'REL-FIN-003': (
        'finance.cash_flow',
        'Fluxo de caixa realizado e projetado',
        'finance',
        'management',
        ('period_start', 'period_end'),
    ),
    'REL-FIN-004': (
        'finance.period_result',
        'Resultado financeiro por período',
        'finance',
        'management',
        ('period_start', 'period_end'),
    ),
    'REL-FIS-001': (
        'fiscal.documents',
        'Documentos fiscais por período e situação',
        'fiscal',
        'operational',
        ('period_start', 'period_end'),
    ),
    'REL-FIS-002': (
        'fiscal.tax_assessment',
        'Apuração de tributos',
        'fiscal',
        'management',
        ('period_start', 'period_end'),
    ),
    'REL-FIS-003': (
        'fiscal.books',
        'Livro de entradas e saídas',
        'fiscal',
        'operational',
        ('period_start', 'period_end'),
    ),
    'REL-EST-001': (
        'inventory.position',
        'Posição de estoque',
        'inventory',
        'operational',
        (),
    ),
    'REL-EST-002': (
        'inventory.expiry',
        'Lotes próximos do vencimento ou vencidos',
        'inventory',
        'operational',
        ('period_end',),
    ),
    'REL-EST-003': (
        'inventory.genealogy',
        'Genealogia e rastreabilidade de lotes',
        'traceability',
        'audit',
        (),
    ),
    'REL-COM-001': (
        'procurement.open_delayed_orders',
        'Pedidos de compra abertos ou atrasados',
        'procurement',
        'operational',
        (),
    ),
    'REL-COM-002': (
        'procurement.receipt_supplier_performance',
        'Divergências de recebimento e fornecedores',
        'procurement',
        'management',
        ('period_start', 'period_end'),
    ),
    'REL-PRO-001': (
        'production.orders_status_delay',
        'Ordens de produção por situação e atraso',
        'production',
        'operational',
        (),
    ),
    'REL-PRO-002': (
        'production.consumption_variance',
        'Consumo planejado versus realizado',
        'production',
        'management',
        ('period_start', 'period_end'),
    ),
    'REL-PRO-003': (
        'production.yield_loss_cost',
        'Rendimento, perdas e custo por ordem',
        'production',
        'management',
        ('period_start', 'period_end'),
    ),
}

CURATED_MODULE_PERMISSIONS = {
    'finance': 'finance.view_financialtitle',
    'fiscal': 'fiscal.view_fiscaldocument',
    'inventory': 'inventory.view_stockbalance',
    'traceability': 'inventory.view_stocklotgenealogy',
    'procurement': 'procurement.view_purchaseorder',
    'production': 'production.view_productionorder',
}


def _declared_executor_allowed_filters(executor, monkeypatch):
    from reports.contracts import ReportContext

    executor_module = importlib.import_module(executor.__module__)

    class AllowedFiltersCaptured(Exception):
        pass

    def capture_allowed_filters(_filters, *, allowed, required=()):
        raise AllowedFiltersCaptured(tuple(allowed))

    monkeypatch.setattr(
        executor_module,
        'normalize_report_filters',
        capture_allowed_filters,
    )
    with pytest.raises(AllowedFiltersCaptured) as captured:
        executor(ReportContext(filters={}, user=None))
    return captured.value.args[0]


def _mutable_json_object_ids(value):
    found = []
    if isinstance(value, (dict, list)):
        found.append(id(value))
        children = value.values() if isinstance(value, dict) else value
        for child in children:
            found.extend(_mutable_json_object_ids(child))
    return found


@pytest.mark.django_db
def test_curated_catalog_contract_matches_executors_permissions_and_model_choices(
    monkeypatch,
):
    from django.contrib.auth.models import Permission

    from reports.catalog import CURATED_REPORTS
    from reports.models import ReportDefinition
    from reports.registry import get_executor

    assert len(CURATED_REPORTS) == 15
    assert {item['code'] for item in CURATED_REPORTS} == set(CURATED_REPORT_EXPECTATIONS)
    assert len({item['executor_key'] for item in CURATED_REPORTS}) == 15
    module_values = set(ReportDefinition.Module.values)
    category_values = set(ReportDefinition.Category.values)

    for item in CURATED_REPORTS:
        expected = CURATED_REPORT_EXPECTATIONS[item['code']]
        executor_key, title, module, category, required_filters = expected
        executor = get_executor(executor_key)
        allowed_filters = _declared_executor_allowed_filters(executor, monkeypatch)

        assert item['executor_key'] == executor_key
        assert item['title'] == title
        assert item['description'] == title
        assert item['module'] == module
        assert item['category'] == category
        assert tuple(item['required_filters']) == required_filters
        assert set(item['filter_schema']) == set(allowed_filters)
        assert set(item['required_filters']) <= set(item['filter_schema'])
        assert item['required_permission'] == CURATED_MODULE_PERMISSIONS[module]
        app_label, codename = item['required_permission'].split('.', 1)
        assert Permission.objects.filter(
            content_type__app_label=app_label,
            codename=codename,
        ).exists()
        assert item['module'] in module_values
        assert item['category'] in category_values
        assert item['allowed_export_formats'] == ['pdf', 'xlsx', 'csv']
        assert item['query_config'] == {'catalog_version': 1}
        assert item['is_active'] is True
        assert item['is_system_managed'] is True


@pytest.mark.django_db
def test_catalog_sync_is_idempotent_and_creates_exactly_fifteen_system_reports():
    from reports.catalog import sync_curated_report_catalog
    from reports.models import ReportDefinition

    sync_curated_report_catalog(ReportDefinition)
    first_ids = dict(ReportDefinition.objects.values_list('code', 'pk'))
    sync_curated_report_catalog(ReportDefinition)

    definitions = ReportDefinition.objects.filter(is_system_managed=True)
    assert definitions.count() == 15
    assert definitions.values('code').distinct().count() == 15
    assert dict(definitions.values_list('code', 'pk')) == first_ids


@pytest.mark.django_db
def test_catalog_sync_preserves_related_history_and_stable_primary_keys():
    from reports.catalog import sync_curated_report_catalog
    from reports.models import ReportDefinition, ReportExecution, ReportSchedule

    sync_curated_report_catalog(ReportDefinition)
    definition = ReportDefinition.objects.get(code='REL-FIN-001')
    original_pk = definition.pk
    schedule = ReportSchedule.objects.create(
        definition=definition,
        name='Agendamento preservado',
        frequency=ReportSchedule.Frequency.DAILY,
        filters={},
        export_format=ReportExecution.ExportFormat.CSV,
        next_run_at=timezone.now(),
    )
    execution = ReportExecution.objects.create(
        definition=definition,
        filters={},
        export_format=ReportExecution.ExportFormat.CSV,
    )
    ReportDefinition.objects.filter(pk=definition.pk).update(
        title='Metadado antigo',
        executor_key='finance.cash_flow',
        query_config={'catalog_version': 0},
    )

    sync_curated_report_catalog(ReportDefinition)

    definition.refresh_from_db()
    schedule.refresh_from_db()
    execution.refresh_from_db()
    assert definition.pk == original_pk
    assert definition.title == CURATED_REPORT_EXPECTATIONS['REL-FIN-001'][1]
    assert definition.executor_key == CURATED_REPORT_EXPECTATIONS['REL-FIN-001'][0]
    assert definition.query_config == {'catalog_version': 1}
    assert schedule.definition_id == original_pk
    assert execution.definition_id == original_pk


@pytest.mark.django_db
def test_curated_catalog_and_persisted_json_have_no_shared_mutable_values():
    from reports.catalog import CURATED_REPORTS, sync_curated_report_catalog
    from reports.models import ReportDefinition

    mutable_ids = [
        object_id for item in CURATED_REPORTS for object_id in _mutable_json_object_ids(item)
    ]
    assert len(mutable_ids) == len(set(mutable_ids))

    sync_curated_report_catalog(ReportDefinition)
    first = ReportDefinition.objects.get(code='REL-FIN-001')
    first.allowed_export_formats.append('mutated-in-memory')
    first.filter_schema['period_start']['label'] = 'Mutado em memória'

    second = ReportDefinition.objects.get(code='REL-FIN-002')
    canonical_first = next(item for item in CURATED_REPORTS if item['code'] == 'REL-FIN-001')
    assert second.allowed_export_formats == ['pdf', 'xlsx', 'csv']
    assert canonical_first['allowed_export_formats'] == ['pdf', 'xlsx', 'csv']
    assert canonical_first['filter_schema']['period_start']['label'] == 'Data inicial'


@pytest.mark.django_db
def test_catalog_sync_rejects_user_definition_collision_before_any_write():
    from reports.catalog import (
        CURATED_REPORT_KEYS,
        CuratedReportCatalogCollision,
        sync_curated_report_catalog,
    )
    from reports.models import ReportDefinition, ReportExecution

    ReportDefinition.objects.filter(code__in=CURATED_REPORT_KEYS).delete()
    collision = ReportDefinition.objects.create(
        code='REL-FIN-003',
        title='Relatório criado pelo usuário',
        module=ReportDefinition.Module.FINANCE,
        category=ReportDefinition.Category.MANAGEMENT,
        allowed_export_formats=[ReportExecution.ExportFormat.CSV],
        query_config={'source': 'legacy'},
        is_system_managed=False,
    )

    with pytest.raises(CuratedReportCatalogCollision) as error:
        sync_curated_report_catalog(ReportDefinition)

    assert str(error.value) == (
        'Código canônico REL-FIN-003 já pertence a uma definição não gerenciada pelo sistema.'
    )
    collision.refresh_from_db()
    assert collision.title == 'Relatório criado pelo usuário'
    assert ReportDefinition.objects.count() == 1
    assert ReportDefinition.objects.filter(is_system_managed=True).exists() is False


@pytest.mark.django_db
def test_catalog_trusted_sync_does_not_weaken_model_or_api_immutability(
    django_user_model,
):
    from rest_framework.test import APIClient

    from reports.catalog import sync_curated_report_catalog
    from reports.models import ReportDefinition

    sync_curated_report_catalog(ReportDefinition)
    definition = ReportDefinition.objects.get(code='REL-FIN-001')
    definition.executor_key = 'finance.cash_flow'
    with pytest.raises(ValidationError):
        definition.save()

    administrator = django_user_model.objects.create_superuser(
        username='catalog-admin@example.com',
        email='catalog-admin@example.com',
        password='StrongPass!123',
    )
    client = APIClient()
    client.force_authenticate(administrator)
    response = client.patch(
        f'/api/reports/definitions/{definition.pk}/',
        {'executor_key': 'finance.cash_flow'},
        format='json',
    )

    assert response.status_code == 400
    assert response.json()['executor_key'] == [
        'Este campo técnico não pode ser alterado em relatório gerenciado pelo sistema.'
    ]


def test_catalog_sync_routes_every_operation_to_the_write_database(monkeypatch):
    from contextlib import contextmanager

    from reports import catalog
    from reports.models import ReportDefinition

    events = []

    class RoutedManager:
        def __init__(self, alias=None):
            self.alias = alias

        @property
        def db(self):
            raise AssertionError('alias de leitura não pode ser consultado')

        def using(self, alias):
            events.append(('using', alias))
            return type(self)(alias)

        def select_for_update(self):
            events.append(('select_for_update', self.alias))
            return self

        def filter(self, **kwargs):
            events.append(('filter', self.alias, kwargs))
            return self

        def create(self, **kwargs):
            events.append(('create', self.alias, kwargs['code']))

        def __iter__(self):
            return iter(())

    class RoutedReportDefinition:
        _meta = ReportDefinition._meta
        _default_manager = RoutedManager('catalog-replica')

    @contextmanager
    def captured_atomic(*, using):
        events.append(('atomic', using))
        yield

    monkeypatch.setattr(
        catalog,
        'router',
        type('WriteRouter', (), {'db_for_write': lambda self, model: 'catalog-primary'})(),
        raising=False,
    )
    monkeypatch.setattr(catalog.transaction, 'atomic', captured_atomic)

    catalog.sync_curated_report_catalog(RoutedReportDefinition)

    assert events[0] == ('using', 'catalog-primary')
    assert ('atomic', 'catalog-primary') in events
    assert sum(event[0] == 'create' for event in events) == 15
    assert {
        event[1] for event in events if event[0] in {'select_for_update', 'filter', 'create'}
    } == {'catalog-primary'}


def test_catalog_sync_retries_whole_transaction_once_then_reraises_integrity_error(
    monkeypatch,
):
    from contextlib import contextmanager

    from django.db import IntegrityError

    from reports import catalog
    from reports.models import ReportDefinition

    attempts = []

    class FailingManager:
        def using(self, alias):
            return self

        def select_for_update(self):
            return self

        def filter(self, **kwargs):
            return self

        def create(self, **kwargs):
            attempts.append(kwargs['code'])
            raise IntegrityError('falha não relacionada persistente')

        def __iter__(self):
            return iter(())

    class RoutedReportDefinition:
        _meta = ReportDefinition._meta
        _default_manager = FailingManager()

    @contextmanager
    def captured_atomic(*, using):
        yield

    monkeypatch.setattr(
        catalog,
        'router',
        type('WriteRouter', (), {'db_for_write': lambda self, model: 'default'})(),
        raising=False,
    )
    monkeypatch.setattr(catalog.transaction, 'atomic', captured_atomic)

    with pytest.raises(IntegrityError, match='falha não relacionada persistente'):
        catalog.sync_curated_report_catalog(RoutedReportDefinition)

    assert attempts == ['REL-FIN-001', 'REL-FIN-001']


@pytest.mark.django_db(transaction=True)
def test_concurrent_empty_catalog_syncs_retry_managed_winner_idempotently(monkeypatch):
    from threading import Lock, get_ident

    from django.db import close_old_connections
    from django.db.models import QuerySet

    from reports.catalog import CURATED_REPORT_KEYS, sync_curated_report_catalog
    from reports.models import ReportDefinition

    ReportDefinition.objects.filter(code__in=CURATED_REPORT_KEYS).delete()
    empty_scan_barrier = Barrier(2)
    synchronized_threads = set()
    synchronized_threads_lock = Lock()
    original_iter = QuerySet.__iter__

    def synchronize_empty_locked_scan(queryset):
        rows = list(original_iter(queryset))
        should_wait = False
        if queryset.model is ReportDefinition and queryset.query.select_for_update and not rows:
            thread_id = get_ident()
            with synchronized_threads_lock:
                if thread_id not in synchronized_threads:
                    synchronized_threads.add(thread_id)
                    should_wait = True
        if should_wait:
            empty_scan_barrier.wait(timeout=10)
        return iter(rows)

    monkeypatch.setattr(QuerySet, '__iter__', synchronize_empty_locked_scan)

    def worker():
        close_old_connections()
        try:
            sync_curated_report_catalog(ReportDefinition)
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=2) as pool:
        futures = [pool.submit(worker) for _index in range(2)]
        for future in futures:
            future.result(timeout=30)

    assert (
        ReportDefinition.objects.filter(
            code__in=CURATED_REPORT_KEYS,
            is_system_managed=True,
        ).count()
        == 15
    )


@pytest.mark.django_db(transaction=True)
def test_empty_catalog_race_with_user_winner_raises_collision_without_partial_writes(
    monkeypatch,
):
    from threading import Event, get_ident

    from django.db import close_old_connections

    from reports.catalog import (
        CURATED_REPORT_KEYS,
        CuratedReportCatalogCollision,
        sync_curated_report_catalog,
    )
    from reports.models import ReportDefinition, ReportExecution

    ReportDefinition.objects.filter(code__in=CURATED_REPORT_KEYS).delete()
    catalog_unique_validation_completed = Event()
    user_winner_committed = Event()
    sync_thread_id = get_ident()
    original_validate_constraints = ReportDefinition.validate_constraints

    def pause_after_catalog_unique_validation(definition, *args, **kwargs):
        result = original_validate_constraints(definition, *args, **kwargs)
        if (
            get_ident() == sync_thread_id
            and definition._state.adding
            and definition.code == 'REL-FIN-003'
            and not catalog_unique_validation_completed.is_set()
        ):
            catalog_unique_validation_completed.set()
            assert user_winner_committed.wait(timeout=10)
        return result

    monkeypatch.setattr(
        ReportDefinition,
        'validate_constraints',
        pause_after_catalog_unique_validation,
    )

    def create_user_winner():
        close_old_connections()
        try:
            assert catalog_unique_validation_completed.wait(timeout=10)
            ReportDefinition.objects.create(
                code='REL-FIN-003',
                title='Definição concorrente do usuário',
                module=ReportDefinition.Module.FINANCE,
                category=ReportDefinition.Category.MANAGEMENT,
                allowed_export_formats=[ReportExecution.ExportFormat.CSV],
                query_config={'source': 'legacy'},
                is_system_managed=False,
            )
            user_winner_committed.set()
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=1) as pool:
        user_winner = pool.submit(create_user_winner)
        with pytest.raises(CuratedReportCatalogCollision):
            sync_curated_report_catalog(ReportDefinition)
        user_winner.result(timeout=20)

    winner = ReportDefinition.objects.get(code='REL-FIN-003')
    assert winner.is_system_managed is False
    assert winner.title == 'Definição concorrente do usuário'
    assert (
        ReportDefinition.objects.filter(
            code__in=CURATED_REPORT_KEYS,
            is_system_managed=True,
        ).exists()
        is False
    )


@pytest.fixture
def restore_latest_report_migrations():
    yield
    from django.db import connection
    from django.db.migrations.executor import MigrationExecutor

    executor = MigrationExecutor(connection)
    executor.migrate(executor.loader.graph.leaf_nodes())


@pytest.mark.django_db(transaction=True)
def test_curated_catalog_migration_replays_without_deleting_or_rekeying_history(
    restore_latest_report_migrations,
):
    from django.db import connection
    from django.db.migrations.executor import MigrationExecutor

    old_target = ('reports', '0004_report_execution_engine')
    new_target = ('reports', '0005_seed_curated_report_catalog')
    executor = MigrationExecutor(connection)
    executor.migrate([old_target])
    old_apps = executor.loader.project_state([old_target]).apps
    OldDefinition = old_apps.get_model('reports', 'ReportDefinition')
    OldExecution = old_apps.get_model('reports', 'ReportExecution')
    OldSchedule = old_apps.get_model('reports', 'ReportSchedule')
    OldDefinition.objects.filter(code__in=CURATED_REPORT_EXPECTATIONS).delete()

    canonical = OldDefinition.objects.create(
        code='REL-FIN-001',
        title='Título canônico legado',
        module='finance',
        category='operational',
        allowed_export_formats=['csv'],
        required_filters=[],
        query_config={'catalog_version': 0},
        executor_key='finance.cash_flow',
        is_system_managed=True,
        filter_schema={},
        required_permission='finance.view_financialtitle',
    )
    legacy = OldDefinition.objects.create(
        code='LEGACY-CUSTOM-001',
        title='Relatório legado preservado',
        module='quality',
        category='operational',
        allowed_export_formats=['csv'],
        required_filters=[],
        query_config={'source': 'quality.QualitySample'},
        is_system_managed=False,
    )
    schedule = OldSchedule.objects.create(
        definition=canonical,
        name='Agenda legada',
        frequency='daily',
        filters={},
        export_format='csv',
        next_run_at=timezone.now(),
    )
    execution = OldExecution.objects.create(
        definition=canonical,
        filters={},
        export_format='csv',
    )

    executor.loader.build_graph()
    executor.migrate([new_target])
    new_apps = executor.loader.project_state([new_target]).apps
    NewDefinition = new_apps.get_model('reports', 'ReportDefinition')
    NewExecution = new_apps.get_model('reports', 'ReportExecution')
    NewSchedule = new_apps.get_model('reports', 'ReportSchedule')

    assert set(
        NewDefinition.objects.filter(
            code__in=CURATED_REPORT_EXPECTATIONS,
            is_system_managed=True,
        ).values_list('code', flat=True)
    ) == set(CURATED_REPORT_EXPECTATIONS)
    migrated = NewDefinition.objects.get(code='REL-FIN-001')
    assert migrated.pk == canonical.pk
    assert migrated.title == CURATED_REPORT_EXPECTATIONS['REL-FIN-001'][1]
    assert NewDefinition.objects.get(pk=legacy.pk).title == 'Relatório legado preservado'
    assert NewSchedule.objects.get(pk=schedule.pk).definition_id == canonical.pk
    assert NewExecution.objects.get(pk=execution.pk).definition_id == canonical.pk

    executor.loader.build_graph()
    executor.migrate([old_target])
    reversed_apps = executor.loader.project_state([old_target]).apps
    ReversedDefinition = reversed_apps.get_model('reports', 'ReportDefinition')
    assert ReversedDefinition.objects.filter(code__in=CURATED_REPORT_EXPECTATIONS).count() == 15
    ReversedDefinition.objects.filter(pk=canonical.pk).update(title='Título após rollback')

    executor.loader.build_graph()
    executor.migrate([new_target])
    replayed_apps = executor.loader.project_state([new_target]).apps
    ReplayedDefinition = replayed_apps.get_model('reports', 'ReportDefinition')
    ReplayedExecution = replayed_apps.get_model('reports', 'ReportExecution')
    ReplayedSchedule = replayed_apps.get_model('reports', 'ReportSchedule')
    replayed = ReplayedDefinition.objects.get(code='REL-FIN-001')
    assert replayed.pk == canonical.pk
    assert replayed.title == CURATED_REPORT_EXPECTATIONS['REL-FIN-001'][1]
    assert ReplayedDefinition.objects.get(pk=legacy.pk).title == 'Relatório legado preservado'
    assert ReplayedSchedule.objects.get(pk=schedule.pk).definition_id == canonical.pk
    assert ReplayedExecution.objects.get(pk=execution.pk).definition_id == canonical.pk


@pytest.mark.django_db(transaction=True)
def test_curated_catalog_migration_rejects_legacy_user_code_collision_atomically(
    restore_latest_report_migrations,
):
    from django.db import connection
    from django.db.migrations.executor import MigrationExecutor
    from django.db.migrations.recorder import MigrationRecorder

    old_target = ('reports', '0004_report_execution_engine')
    new_target = ('reports', '0005_seed_curated_report_catalog')
    executor = MigrationExecutor(connection)
    executor.migrate([old_target])
    old_apps = executor.loader.project_state([old_target]).apps
    OldDefinition = old_apps.get_model('reports', 'ReportDefinition')
    OldDefinition.objects.filter(code__in=CURATED_REPORT_EXPECTATIONS).delete()
    collision = OldDefinition.objects.create(
        code='REL-FIN-003',
        title='Definição do usuário',
        module='finance',
        category='management',
        allowed_export_formats=['csv'],
        required_filters=[],
        query_config={'source': 'legacy'},
        is_system_managed=False,
    )

    executor.loader.build_graph()
    with pytest.raises(
        RuntimeError,
        match=(
            'Código canônico REL-FIN-003 já pertence a uma definição não gerenciada pelo sistema'
        ),
    ):
        executor.migrate([new_target])

    collision.refresh_from_db()
    assert collision.title == 'Definição do usuário'
    assert OldDefinition.objects.count() == 1
    assert (
        MigrationRecorder(connection)
        .migration_qs.filter(
            app='reports',
            name='0005_seed_curated_report_catalog',
        )
        .exists()
        is False
    )
    collision.delete()


@pytest.mark.django_db
def test_run_report_serializer_accepts_only_operational_inputs_and_validates_definition(
    system_report_definition,
):
    from reports.serializers import RunReportSerializer

    serializer = RunReportSerializer(
        data={
            'export_format': 'csv',
            'filters': {},
            'executor_key': 'attacker.executor',
            'query_config': {'source': 'auth.User'},
            'filter_schema': {'secret': {'type': 'text'}},
            'requested_by': 999999,
            'status': 'completed',
            'result_reference': '/tmp/forged.csv',
        },
        context={'definition': system_report_definition},
    )

    assert serializer.is_valid(), serializer.errors
    assert serializer.validated_data == {'export_format': 'csv', 'filters': {}}
    invalid = RunReportSerializer(
        data={'export_format': 'csv', 'filters': {'unsupported_key': 'secret'}},
        context={'definition': system_report_definition},
    )
    assert invalid.is_valid() is False
    assert 'filters' in invalid.errors


@pytest.mark.parametrize(
    'filters',
    (
        [],
        [('status', 'open')],
        'status=open',
        42,
        {'status': float('nan')},
    ),
)
@pytest.mark.django_db
def test_run_report_serializer_rejects_nonobject_or_unsafe_filter_payloads(
    system_report_definition,
    filters,
):
    from reports.serializers import RunReportSerializer

    serializer = RunReportSerializer(
        data={'export_format': 'csv', 'filters': filters},
        context={'definition': system_report_definition},
    )

    assert serializer.is_valid() is False
    assert 'filters' in serializer.errors


@pytest.mark.parametrize('filters', ([], '', 0, [('status', 'open')]))
@pytest.mark.django_db
def test_report_definition_domain_rejects_nonobject_filters(
    system_report_definition,
    filters,
):
    with pytest.raises(ValidationError) as invalid:
        system_report_definition.normalize_filters(filters)

    assert invalid.value.message_dict == {
        'filters': ['Filtros devem ser um objeto JSON seguro.'],
    }


@pytest.mark.django_db
def test_report_run_api_uses_view_add_and_domain_permissions_without_change_permission(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution

    _configure_report_encryption(settings, tmp_path)
    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='view_reportdefinition',
        )
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.post(
        f'/api/reports/definitions/{system_report_definition.pk}/run/',
        {
            'export_format': 'csv',
            'filters': {},
            'executor_key': 'attacker.executor',
            'query_config': {'source': 'auth.User'},
            'requested_by': system_report_definition.owner_id,
            'status': 'completed',
        },
        format='json',
    )

    assert response.status_code == 201
    execution = ReportExecution.objects.get(pk=response.json()['id'])
    assert execution.status == ReportExecution.Status.COMPLETED
    assert execution.requested_by == report_execution_actor
    assert execution.definition == system_report_definition
    system_report_definition.refresh_from_db()
    assert system_report_definition.executor_key == 'tests.lifecycle'
    assert report_execution_actor.has_perm('reports.change_reportdefinition') is False


@pytest.mark.django_db
def test_report_run_api_denies_missing_domain_permission_before_creating_execution(
    system_report_definition,
    django_user_model,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution

    user = django_user_model.objects.create_user(
        username='report.no-domain@example.com',
        email='report.no-domain@example.com',
    )
    user.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label='reports',
            codename__in=('view_reportdefinition', 'add_reportexecution'),
        )
    )
    client = APIClient()
    client.force_authenticate(user)
    count_before = ReportExecution.objects.count()

    response = client.post(
        f'/api/reports/definitions/{system_report_definition.pk}/run/',
        {'export_format': 'csv', 'filters': {}},
        format='json',
    )

    assert response.status_code == 403
    assert ReportExecution.objects.count() == count_before


@pytest.mark.django_db
def test_raw_report_definition_api_requires_change_while_run_does_not(
    system_report_definition,
    report_execution_actor,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='view_reportdefinition',
        )
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    list_response = client.get('/api/reports/definitions/')
    detail_response = client.get(f'/api/reports/definitions/{system_report_definition.pk}/')

    assert list_response.status_code == 403
    assert detail_response.status_code == 403


@pytest.mark.django_db
def test_raw_definition_create_and_destroy_require_change_in_addition_to_model_action(
    system_report_definition,
    django_user_model,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportDefinition

    user = django_user_model.objects.create_user(
        username='report.raw-actions@example.com',
        email='report.raw-actions@example.com',
    )
    user.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label='reports',
            codename__in=('add_reportdefinition', 'delete_reportdefinition'),
        )
    )
    client = APIClient()
    client.force_authenticate(user)
    count_before = ReportDefinition.objects.count()

    create_response = client.post(
        '/api/reports/definitions/',
        {
            'code': 'RAW-DENIED',
            'title': 'Definição crua negada',
            'module': 'finance',
            'category': 'operational',
            'allowed_export_formats': ['csv'],
        },
        format='json',
    )
    destroy_response = client.delete(f'/api/reports/definitions/{system_report_definition.pk}/')

    assert create_response.status_code == 403
    assert destroy_response.status_code == 403
    assert ReportDefinition.objects.count() == count_before
    assert ReportDefinition.objects.filter(pk=system_report_definition.pk).exists()


@pytest.mark.django_db
def test_explicit_action_permission_map_is_enforced_for_get_and_post(
    django_user_model,
):
    from types import SimpleNamespace

    from django.contrib.auth.models import Permission

    from base.permissions import SingleInstanceDjangoModelPermissions
    from reports.models import ReportDefinition

    user = django_user_model.objects.create_user(
        username='report.action-map@example.com',
        email='report.action-map@example.com',
    )
    permission = SingleInstanceDjangoModelPermissions()
    view = SimpleNamespace(
        action='download',
        queryset=ReportDefinition.objects.none(),
        action_permission_map={
            'download': (
                'reports.view_reportexecution',
                'files.view_protectedfile',
            ),
            'run': (
                'reports.view_reportdefinition',
                'reports.add_reportexecution',
            ),
        },
    )
    get_request = SimpleNamespace(method='GET', user=user)
    post_request = SimpleNamespace(method='POST', user=user)

    def get_permission(app_label, codename):
        return Permission.objects.get(
            content_type__app_label=app_label,
            codename=codename,
        )

    user.user_permissions.add(get_permission('reports', 'view_reportexecution'))
    assert permission.has_permission(get_request, view) is False
    user.user_permissions.add(get_permission('files', 'view_protectedfile'))
    for cache_name in ('_perm_cache', '_user_perm_cache', '_group_perm_cache'):
        if hasattr(user, cache_name):
            delattr(user, cache_name)
    assert permission.has_permission(get_request, view) is True

    view.action = 'run'
    user.user_permissions.add(get_permission('reports', 'view_reportdefinition'))
    for cache_name in ('_perm_cache', '_user_perm_cache', '_group_perm_cache'):
        if hasattr(user, cache_name):
            delattr(user, cache_name)
    assert permission.has_permission(post_request, view) is False
    user.user_permissions.add(get_permission('reports', 'add_reportexecution'))
    for cache_name in ('_perm_cache', '_user_perm_cache', '_group_perm_cache'):
        if hasattr(user, cache_name):
            delattr(user, cache_name)
    assert permission.has_permission(post_request, view) is True


@pytest.mark.parametrize('filters', ([], 'status=open', 7))
@pytest.mark.django_db
def test_report_run_api_returns_400_for_nonobject_filters_without_orphan_execution(
    system_report_definition,
    report_execution_actor,
    filters,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution

    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='view_reportdefinition',
        )
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)
    count_before = ReportExecution.objects.count()

    response = client.post(
        f'/api/reports/definitions/{system_report_definition.pk}/run/',
        {'export_format': 'csv', 'filters': filters},
        format='json',
    )

    assert response.status_code == 400
    assert 'filters' in response.json()
    assert ReportExecution.objects.count() == count_before


@pytest.mark.django_db
def test_completed_report_download_uses_shared_aes_path_safe_headers_and_audit(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from files.models import ProtectedFile, ProtectedFileAuditTrail

    _configure_report_encryption(settings, tmp_path)
    report_execution_actor.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label__in=('reports', 'files'),
            codename__in=('view_reportexecution', 'view_protectedfile'),
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    ProtectedFile.objects.filter(pk=execution.result_file_id).update(
        file_name='../../resultado\r\nX-Injected: yes.csv',
    )
    execution.result_file.refresh_from_db()
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    report_response = client.get(
        f'/api/reports/executions/{execution.pk}/download/',
        REMOTE_ADDR='198.51.100.24',
        HTTP_USER_AGENT='Task6 browser',
    )
    file_response = client.get(
        f'/api/files/protected-files/{execution.result_file_id}/download/',
        REMOTE_ADDR='198.51.100.24',
        HTTP_USER_AGENT='Task6 browser',
    )

    assert report_response.status_code == 200
    assert report_response.content.startswith(b'\xef\xbb\xbf')
    assert report_response['Content-Type'].startswith('text/csv')
    assert '\r' not in report_response['Content-Disposition']
    assert '\n' not in report_response['Content-Disposition']
    assert '../' not in report_response['Content-Disposition']
    assert file_response.status_code == 200
    assert file_response.content == report_response.content
    audits = ProtectedFileAuditTrail.objects.filter(
        protected_file=execution.result_file,
        action=ProtectedFileAuditTrail.Action.DOWNLOAD,
    ).order_by('pk')
    assert audits.count() == 2
    assert {audit.ip_address for audit in audits} == {'198.51.100.24'}
    assert {audit.user_agent for audit in audits} == {'Task6 browser'}


@pytest.mark.parametrize('failure_mode', ('missing', 'corrupt'))
@pytest.mark.django_db
def test_report_download_storage_or_cipher_failure_is_generic_and_never_audits_download(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    failure_mode,
):
    from django.contrib.auth.models import Permission
    from django.core.files.storage import default_storage
    from rest_framework.test import APIClient

    from files.models import ProtectedFileAuditTrail

    _configure_report_encryption(settings, tmp_path)
    report_execution_actor.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label__in=('reports', 'files'),
            codename__in=('view_reportexecution', 'view_protectedfile'),
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    if failure_mode == 'missing':
        default_storage.delete(execution.result_file.file_reference)
    else:
        (tmp_path / execution.result_file.file_reference).write_bytes(b'not-an-aes-envelope')
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.get(
        f'/api/reports/executions/{execution.pk}/download/',
        REMOTE_ADDR='192.0.2.19',
        HTTP_USER_AGENT='Broken storage browser',
    )

    assert response.status_code in (403, 404)
    assert response.json() == {'detail': 'Arquivo de relatório indisponível.'}
    assert (
        ProtectedFileAuditTrail.objects.filter(
            protected_file=execution.result_file,
            action=ProtectedFileAuditTrail.Action.DOWNLOAD,
        ).exists()
        is False
    )
    denied = ProtectedFileAuditTrail.objects.get(
        protected_file=execution.result_file,
        action=ProtectedFileAuditTrail.Action.ACCESS_DENIED,
    )
    assert denied.ip_address == '192.0.2.19'
    assert denied.user_agent == 'Broken storage browser'
    assert denied.details == {'reason': 'content_unavailable'}


@pytest.mark.django_db
def test_report_download_denies_file_access_rule_even_with_both_django_permissions(
    system_report_definition,
    report_execution_actor,
    django_user_model,
    settings,
    tmp_path,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from files.models import ProtectedFileAuditTrail

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    outsider = django_user_model.objects.create_user(
        username='report.outsider@example.com',
        email='report.outsider@example.com',
    )
    outsider.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label__in=('reports', 'files'),
            codename__in=('view_reportexecution', 'view_protectedfile'),
        )
    )
    client = APIClient()
    client.force_authenticate(outsider)

    response = client.get(
        f'/api/reports/executions/{execution.pk}/download/',
        REMOTE_ADDR='203.0.113.9',
        HTTP_USER_AGENT='Denied browser',
    )

    assert response.status_code == 403
    assert execution.result_file.file_reference.encode() not in response.content
    denied = ProtectedFileAuditTrail.objects.get(
        protected_file=execution.result_file,
        action=ProtectedFileAuditTrail.Action.ACCESS_DENIED,
        actor=outsider,
    )
    assert denied.ip_address == '203.0.113.9'
    assert denied.user_agent == 'Denied browser'


@pytest.mark.parametrize('execution_status', ('pending', 'running', 'failed', 'cancelled'))
@pytest.mark.django_db
def test_report_download_never_returns_noncompleted_artifact(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    execution_status,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution

    _configure_report_encryption(settings, tmp_path)
    report_execution_actor.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label__in=('reports', 'files'),
            codename__in=('view_reportexecution', 'view_protectedfile'),
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    ReportExecution.objects.filter(pk=execution.pk).update(status=execution_status)
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.get(f'/api/reports/executions/{execution.pk}/download/')

    assert response.status_code in (403, 404, 409)
    assert execution.result_file.file_reference.encode() not in response.content


@pytest.mark.django_db
def test_completed_report_without_artifact_returns_generic_not_found(
    system_report_definition,
    report_execution_actor,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportExecution

    report_execution_actor.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label__in=('reports', 'files'),
            codename__in=('view_reportexecution', 'view_protectedfile'),
        )
    )
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    ReportExecution.objects.filter(pk=execution.pk).update(
        status=ReportExecution.Status.COMPLETED,
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.get(f'/api/reports/executions/{execution.pk}/download/')

    assert response.status_code == 404
    assert response.json() == {'detail': 'Arquivo de relatório indisponível.'}


@pytest.mark.django_db
def test_report_download_requires_both_execution_and_file_model_permissions(
    system_report_definition,
    report_execution_actor,
    django_user_model,
    settings,
    tmp_path,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    _configure_report_encryption(settings, tmp_path)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    for codename in ('view_reportexecution', 'view_protectedfile'):
        user = django_user_model.objects.create_user(
            username=f'{codename}@example.com',
            email=f'{codename}@example.com',
        )
        other_codename = (
            'view_protectedfile' if codename == 'view_reportexecution' else 'view_reportexecution'
        )
        user.user_permissions.add(Permission.objects.get(codename=other_codename))
        client = APIClient()
        client.force_authenticate(user)

        response = client.get(f'/api/reports/executions/{execution.pk}/download/')

        assert response.status_code == 403
        assert execution.result_file.file_reference.encode() not in response.content


SYSTEM_FILTER_SCHEMA = {
    'period_start': {'type': 'date', 'label': 'Data inicial'},
    'status': {
        'type': 'choice',
        'label': 'Situação',
        'choices': [
            {'value': 'open', 'label': 'Em aberto'},
            {'value': 'closed', 'label': 'Encerrado'},
        ],
    },
    'lot': {'type': 'text', 'label': 'Lote'},
    'customer': {'type': 'integer', 'label': 'Cliente'},
}


@pytest.mark.parametrize(
    'filters',
    (
        {'period_start': '2026-07-01', 'status': 'open', 'supplier': 1},
        {'period_start': 20260701, 'status': 'open'},
        {'period_start': '2026-07-01', 'status': 'forged'},
        {'period_start': '2026-07-01', 'status': {'value': 'open'}},
        {'period_start': '2026-07-01', 'status': 'open', 'customer': '7'},
        {'period_start': '2026-07-01'},
    ),
)
@pytest.mark.django_db
def test_system_managed_definition_rejects_schema_mismatch_before_execution_create(
    system_report_definition,
    report_execution_actor,
    filters,
):
    from reports.models import ReportDefinition, ReportExecution

    ReportDefinition.objects.filter(pk=system_report_definition.pk).update(
        filter_schema=SYSTEM_FILTER_SCHEMA,
        required_filters=['period_start', 'status'],
    )
    system_report_definition.refresh_from_db()
    count_before = ReportExecution.objects.count()

    with pytest.raises(ValidationError):
        system_report_definition.create_execution(
            filters=filters,
            export_format='csv',
            requested_by=report_execution_actor,
        )

    assert ReportExecution.objects.count() == count_before


@pytest.mark.django_db
def test_system_managed_definition_accepts_exact_typed_schema_payload(
    system_report_definition,
    report_execution_actor,
):
    from reports.models import ReportDefinition

    ReportDefinition.objects.filter(pk=system_report_definition.pk).update(
        filter_schema=SYSTEM_FILTER_SCHEMA,
        required_filters=['period_start', 'status'],
    )
    system_report_definition.refresh_from_db()

    execution = system_report_definition.create_execution(
        filters={
            'period_start': '2026-07-01',
            'status': 'open',
            'lot': 'LOT-001',
            'customer': 7,
        },
        export_format='csv',
        requested_by=report_execution_actor,
    )

    assert execution.filters == {
        'period_start': '2026-07-01',
        'status': 'open',
        'lot': 'LOT-001',
        'customer': 7,
    }


@pytest.mark.parametrize('corrupted_defaults', ([], '', None))
@pytest.mark.django_db
def test_system_managed_definition_fails_closed_for_non_object_default_filters(
    system_report_definition,
    report_execution_actor,
    corrupted_defaults,
):
    from reports.models import ReportExecution

    system_report_definition.default_filters = corrupted_defaults
    count_before = ReportExecution.objects.count()

    with pytest.raises(ValidationError):
        system_report_definition.create_execution(
            filters={},
            export_format='csv',
            requested_by=report_execution_actor,
        )

    assert ReportExecution.objects.count() == count_before


@pytest.mark.django_db
def test_non_system_definition_keeps_explicit_legacy_global_filter_fallback(
    django_user_model,
):
    from reports.models import ReportDefinition

    owner = django_user_model.objects.create_user(
        username='legacy.filters@example.com',
        email='legacy.filters@example.com',
    )
    definition = ReportDefinition.objects.create(
        code='LEGACY-FILTER-FALLBACK',
        title='Relatório legado',
        module='finance',
        category='operational',
        allowed_export_formats=['csv'],
        filter_schema={},
        owner=owner,
        is_system_managed=False,
    )

    assert definition.normalize_filters({'supplier': 'SUP-01', 'product': 'PRD-01'}) == {
        'supplier': 'SUP-01',
        'product': 'PRD-01',
    }


@pytest.mark.django_db
def test_run_report_definition_reloads_locked_active_definition_before_create(
    system_report_definition,
    report_execution_actor,
):
    from reports.models import ReportDefinition, ReportExecution
    from reports.services import run_report_definition

    stale_definition = system_report_definition
    ReportDefinition.objects.filter(pk=stale_definition.pk).update(is_active=False)
    count_before = ReportExecution.objects.count()

    with pytest.raises(ValidationError) as inactive:
        run_report_definition(
            definition=stale_definition,
            user=report_execution_actor,
            filters={},
            export_format='csv',
        )

    assert inactive.value.message_dict == {'definition': ['O relatório está inativo.']}
    assert ReportExecution.objects.count() == count_before


@pytest.mark.django_db
def test_run_report_api_rejects_inactive_definition_without_pending_row(
    system_report_definition,
    report_execution_actor,
):
    from django.contrib.auth.models import Permission
    from rest_framework.test import APIClient

    from reports.models import ReportDefinition, ReportExecution

    report_execution_actor.user_permissions.add(
        Permission.objects.get(
            content_type__app_label='reports',
            codename='view_reportdefinition',
        )
    )
    ReportDefinition.objects.filter(pk=system_report_definition.pk).update(is_active=False)
    client = APIClient()
    client.force_authenticate(report_execution_actor)
    count_before = ReportExecution.objects.count()

    response = client.post(
        f'/api/reports/definitions/{system_report_definition.pk}/run/',
        {'export_format': 'csv', 'filters': {}},
        format='json',
    )

    assert response.status_code == 400
    assert response.json() == {'definition': ['O relatório está inativo.']}
    assert ReportExecution.objects.count() == count_before


def _grant_report_download_permissions(user):
    from django.contrib.auth.models import Permission

    user.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label__in=('reports', 'files'),
            codename__in=(
                'view_reportexecution',
                'view_protectedfile',
                'change_securefilelink',
            ),
        )
    )


@pytest.mark.parametrize('valid_until_mode', ('future', 'none'))
@pytest.mark.django_db
def test_expired_status_is_denied_even_when_valid_until_has_not_expired(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
    valid_until_mode,
):
    from datetime import timedelta

    from django.utils import timezone
    from rest_framework.test import APIClient

    from files.models import ProtectedFile, ProtectedFileAuditTrail

    _configure_report_encryption(settings, tmp_path)
    _grant_report_download_permissions(report_execution_actor)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    valid_until = (
        timezone.localdate() + timedelta(days=30) if valid_until_mode == 'future' else None
    )
    ProtectedFile.objects.filter(pk=execution.result_file_id).update(
        status=ProtectedFile.Status.EXPIRED,
        valid_until=valid_until,
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.get(
        f'/api/reports/executions/{execution.pk}/download/',
        REMOTE_ADDR='192.0.2.44',
        HTTP_USER_AGENT='Expired status browser',
    )

    assert response.status_code == 403
    assert response.json() == {'detail': 'Arquivo de relatório indisponível.'}
    denied = ProtectedFileAuditTrail.objects.get(
        protected_file_id=execution.result_file_id,
        action=ProtectedFileAuditTrail.Action.ACCESS_DENIED,
    )
    assert denied.ip_address == '192.0.2.44'
    assert denied.user_agent == 'Expired status browser'


@pytest.mark.django_db
def test_external_serializers_and_secure_link_use_never_expose_storage_references(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from rest_framework.test import APIClient

    from files.models import ProtectedFileAuditTrail, SecureFileLink
    from files.serializers import (
        ProtectedFileAuditTrailSerializer,
        ProtectedFileSerializer,
    )
    from reports.serializers import ReportExecutionSerializer

    _configure_report_encryption(settings, tmp_path)
    _grant_report_download_permissions(report_execution_actor)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    internal_reference = execution.result_file.file_reference
    link = execution.result_file.generate_secure_link(
        report_execution_actor,
        purpose=SecureFileLink.Purpose.DOWNLOAD,
    )
    audit = execution.result_file.record_audit(
        ProtectedFileAuditTrail.Action.VIEW,
        user=report_execution_actor,
        details={
            'file_reference': internal_reference,
            'nested': {'storage_path': internal_reference, 'safe': 'preserved'},
        },
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    execution_response = client.get(f'/api/reports/executions/{execution.pk}/')
    file_response = client.get(f'/api/files/protected-files/{execution.result_file_id}/')
    use_response = client.post(f'/api/files/secure-links/{link.pk}/use/')

    assert 'result_reference' not in ReportExecutionSerializer(execution).data
    assert 'file_reference' not in ProtectedFileSerializer(execution.result_file).data
    assert ProtectedFileAuditTrailSerializer(audit).data['details'] == {
        'nested': {'safe': 'preserved'}
    }
    assert 'result_reference' not in execution_response.json()
    assert 'file_reference' not in file_response.json()
    assert 'file_reference' not in use_response.json()
    assert internal_reference not in execution_response.content.decode()
    assert internal_reference not in file_response.content.decode()
    assert internal_reference not in use_response.content.decode()


@pytest.mark.django_db
def test_both_download_routes_set_private_no_store_headers_and_reject_crlf_mime(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from rest_framework.test import APIClient

    from files.models import ProtectedFile

    _configure_report_encryption(settings, tmp_path)
    _grant_report_download_permissions(report_execution_actor)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    ProtectedFile.objects.filter(pk=execution.result_file_id).update(
        mime_type='text/csv;\r\ncharset=utf-8',
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    responses = (
        client.get(f'/api/reports/executions/{execution.pk}/download/'),
        client.get(f'/api/files/protected-files/{execution.result_file_id}/download/'),
    )

    for response in responses:
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/octet-stream'
        assert response['Cache-Control'] == 'private, no-store, max-age=0'
        assert response['Pragma'] == 'no-cache'
        assert response['Expires'] == '0'
        assert {'authorization', 'cookie'} <= {
            value.strip().casefold() for value in response['Vary'].split(',')
        }


@pytest.mark.django_db
def test_well_formed_cipher_envelope_with_invalid_nonce_is_generic_and_audited(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from rest_framework.test import APIClient

    from files.models import ProtectedFileAuditTrail

    _configure_report_encryption(settings, tmp_path)
    _grant_report_download_permissions(report_execution_actor)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    (tmp_path / execution.result_file.file_reference).write_bytes(b'aes256gcm:v1:test:AA:AA')
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.get(
        f'/api/reports/executions/{execution.pk}/download/',
        REMOTE_ADDR='198.51.100.61',
        HTTP_USER_AGENT='Invalid nonce browser',
    )

    assert response.status_code == 403
    assert response.json() == {'detail': 'Arquivo de relatório indisponível.'}
    assert (
        ProtectedFileAuditTrail.objects.filter(
            protected_file=execution.result_file,
            action=ProtectedFileAuditTrail.Action.DOWNLOAD,
        ).exists()
        is False
    )
    denied = ProtectedFileAuditTrail.objects.get(
        protected_file=execution.result_file,
        action=ProtectedFileAuditTrail.Action.ACCESS_DENIED,
    )
    assert denied.details == {'reason': 'content_unavailable'}


@pytest.mark.django_db
def test_nonterminal_report_with_artifact_records_generic_access_denied_metadata(
    system_report_definition,
    report_execution_actor,
    settings,
    tmp_path,
):
    from rest_framework.test import APIClient

    from files.models import ProtectedFileAuditTrail
    from reports.models import ReportExecution

    _configure_report_encryption(settings, tmp_path)
    _grant_report_download_permissions(report_execution_actor)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    ).run(user=report_execution_actor)
    ReportExecution.objects.filter(pk=execution.pk).update(
        status=ReportExecution.Status.PENDING,
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)

    response = client.get(
        f'/api/reports/executions/{execution.pk}/download/',
        REMOTE_ADDR='203.0.113.71',
        HTTP_USER_AGENT='Nonterminal browser',
    )

    assert response.status_code == 404
    assert response.json() == {'detail': 'Arquivo de relatório indisponível.'}
    denied = ProtectedFileAuditTrail.objects.get(
        protected_file=execution.result_file,
        action=ProtectedFileAuditTrail.Action.ACCESS_DENIED,
    )
    assert denied.ip_address == '203.0.113.71'
    assert denied.user_agent == 'Nonterminal browser'
    assert denied.details == {'reason': 'execution_not_completed'}


@pytest.mark.django_db
def test_nonterminal_report_without_artifact_does_not_create_file_audit(
    system_report_definition,
    report_execution_actor,
):
    from rest_framework.test import APIClient

    from files.models import ProtectedFileAuditTrail

    _grant_report_download_permissions(report_execution_actor)
    execution = system_report_definition.create_execution(
        filters={},
        export_format='csv',
        requested_by=report_execution_actor,
    )
    client = APIClient()
    client.force_authenticate(report_execution_actor)
    count_before = ProtectedFileAuditTrail.objects.count()

    response = client.get(f'/api/reports/executions/{execution.pk}/download/')

    assert response.status_code == 404
    assert response.json() == {'detail': 'Arquivo de relatório indisponível.'}
    assert ProtectedFileAuditTrail.objects.count() == count_before


@pytest.mark.parametrize(
    ('configured_value', 'submitted_value'),
    (
        ('open', 'open'),
        (7, '7'),
    ),
)
@pytest.mark.django_db
def test_report_run_form_choice_renders_empty_option_and_preserves_configured_type(
    system_report_definition,
    configured_value,
    submitted_value,
):
    from reports.forms import ReportRunForm

    system_report_definition.filter_schema = {
        'status': {
            'type': 'choice',
            'label': 'Situação',
            'choices': [{'value': configured_value, 'label': 'Configurada'}],
        }
    }
    system_report_definition.required_filters = []

    unbound_form = ReportRunForm(definition=system_report_definition)
    bound_form = ReportRunForm(
        data={'export_format': 'csv', 'status': submitted_value},
        definition=system_report_definition,
    )

    assert list(unbound_form.fields['status'].choices) == [
        ('', '---------'),
        (str(configured_value), 'Configurada'),
    ]
    assert '<option value="" selected>---------</option>' in str(unbound_form['status'])
    assert bound_form.is_valid(), bound_form.errors
    assert bound_form.cleaned_filters == {'status': configured_value}
    assert type(bound_form.cleaned_filters['status']) is type(configured_value)


@pytest.mark.django_db
def test_report_run_form_required_choice_rejects_empty_selection(
    system_report_definition,
):
    from reports.forms import ReportRunForm

    system_report_definition.filter_schema = {
        'status': {
            'type': 'choice',
            'label': 'Situação',
            'choices': [{'value': 'open', 'label': 'Em aberto'}],
        }
    }
    system_report_definition.required_filters = ['status']

    form = ReportRunForm(
        data={'export_format': 'csv', 'status': ''},
        definition=system_report_definition,
    )

    assert form.is_valid() is False
    assert form.errors.as_data()['status'][0].code == 'required'
    assert form.cleaned_filters == {}


@pytest.mark.django_db
def test_report_run_form_optional_choice_omits_empty_selection_from_filters(
    system_report_definition,
):
    from reports.forms import ReportRunForm

    system_report_definition.filter_schema = {
        'status': {
            'type': 'choice',
            'label': 'Situação',
            'choices': [{'value': 7, 'label': 'Configurada'}],
        }
    }
    system_report_definition.required_filters = []

    form = ReportRunForm(
        data={'export_format': 'csv', 'status': ''},
        definition=system_report_definition,
    )

    assert form.is_valid(), form.errors
    assert form.cleaned_data['status'] == ''
    assert form.cleaned_filters == {}
